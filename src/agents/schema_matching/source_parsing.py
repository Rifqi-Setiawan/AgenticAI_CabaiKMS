"""Shared source-spreadsheet parsing helpers — pulled out of
eval/review_schema_matching.py so both that harness and src/ui/ can reuse
the exact same parsing logic rather than maintaining two copies.

Pure I/O + structural parsing, no agent calls: turns a raw source workbook
into a flat list of ParsedAttribute (one per source column/row, with its
structural context and values), for the schema-matching agents to then
retrieve/rerank/normalize against.

`row_values` is positionally aligned across every ParsedAttribute returned
by the SAME call (index i means the same source row for row-oriented, or
the same variety column for transposed) — this is what lets a caller
reconstruct "this specific value belongs to this specific varietas" once
it knows which position maps to which varietas name (the anchor column's
own row_values, for row-oriented; `variety_names` returned directly by
load_transposed_rows, for transposed). `sample_values` (deduplication-free,
None-filtered) stays around for retrieval/rerank prompting, which never
cared about position, only "what values look like".
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl


@dataclass
class ParsedAttribute:
    attribute_name: str
    structural_context: str | None
    row_values: list[str | None]

    @property
    def sample_values(self) -> list[str]:
        return [v for v in self.row_values if v is not None]

    @property
    def display_name(self) -> str:
        """Unambiguous human label without changing the LLM-facing name.

        A multilevel sheet may legitimately contain the same leaf header in
        different sections, for example Young Fruit / Fruit Length and
        Mature Fruit / Fruit Length.
        """
        if self.structural_context:
            return f"{self.structural_context} / {self.attribute_name}"
        return self.attribute_name


def load_row_oriented_columns(
    path: Path, sheet_name: str, *, header_rows: int | None = None,
) -> list[ParsedAttribute]:
    """Read flat one-row or hierarchical two-row headers without losing
    the first observation. Auto mode recognizes merged headers in rows 1–2;
    otherwise a complete first row is a flat header. Ambiguous sparse
    headers require an explicit header_rows=1 or 2 rather than guessing.
    Unmerged two-row headers can always be selected explicitly.

    Every returned attribute has the same observation positions, including
    empty values. Two-row parsing retains the existing forward-filled
    section context and standalone row-1 header behavior.
    """
    if header_rows not in (None, 1, 2):
        raise ValueError("Jumlah baris header harus 1, 2, atau otomatis (None).")
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        merged_headers = [
            r for r in ws.merged_cells.ranges
            if r.min_row == 1 and r.max_row <= 2
        ]
    finally:
        wb.close()

    def present(value: object) -> bool:
        return value is not None and str(value).strip() != ""

    if not rows or not any(present(v) for v in rows[0]):
        raise ValueError("Header kosong. Tempatkan nama kolom pada baris pertama sheet.")
    if header_rows is None:
        # A horizontal merge alone may be a title, not a hierarchy. Require
        # subheaders beneath it, or a standalone header merged vertically.
        hierarchical = len(rows) > 1 and any(
            r.max_row == 2 or (
                r.max_col > r.min_col
                and any(present(v) for v in rows[1][r.min_col - 1:r.max_col])
            ) for r in merged_headers
        )
        header_rows = 2 if hierarchical else 1
        if not hierarchical and len(rows) > 1 and any(
            not present(header) and present(value)
            for header, value in zip(rows[0], rows[1])
        ):
            raise ValueError(
                "Struktur header ambigu. Pilih jumlah baris header 1 atau 2 "
                "sesuai file, dan pastikan setiap kolom data memiliki nama."
            )
    if len(rows) <= header_rows or not any(
        present(v) for row in rows[header_rows:] for v in row
    ):
        raise ValueError("Tidak ada baris data setelah header.")

    if header_rows == 1:
        attributes = []
        for col_idx, header in enumerate(rows[0]):
            if not present(header):
                if any(present(row[col_idx]) for row in rows[1:]):
                    raise ValueError(f"Kolom {col_idx + 1} berisi data tetapi header kosong.")
                continue
            attributes.append(ParsedAttribute(
                str(header).strip(), None,
                [str(row[col_idx]) if present(row[col_idx]) else None for row in rows[1:]],
            ))
        _validate_attribute_names(attributes)
        return attributes

    section_row, field_row, data_rows = rows[0], rows[1], rows[2:]

    sections: list[str | None] = []
    current: str | None = None
    for cell in section_row:
        if present(cell):
            current = str(cell).strip()
        sections.append(current)

    n_cols = max(len(section_row), len(field_row))
    attributes = []
    for col_idx in range(n_cols):
        filled_section = sections[col_idx] if col_idx < len(sections) else None
        raw_section = section_row[col_idx] if col_idx < len(section_row) else None
        field = field_row[col_idx] if col_idx < len(field_row) else None

        if present(field):
            attribute_name, structural_context = str(field).strip(), filled_section
        elif present(raw_section):
            # Only a column that genuinely HAS its own row-1 text (not one
            # merely inheriting a forward-filled value from an earlier
            # column) counts as a standalone, header-only attribute — this
            # is what keeps trailing empty columns (see docs/PROFILING.md
            # §2.2: cols O-Z are unused) from being misread as more
            # instances of whatever section header came last.
            attribute_name, structural_context = str(raw_section).strip(), None
        else:
            continue  # fully empty column

        row_values = [
            (str(row[col_idx]) if col_idx < len(row) and row[col_idx] is not None else None)
            for row in data_rows
        ]
        attributes.append(ParsedAttribute(attribute_name, structural_context, row_values))
    _validate_attribute_names(attributes)
    return attributes


def _validate_attribute_names(attributes: list[ParsedAttribute]) -> None:
    # A repeated leaf header is valid under a different parent section.
    # Only the full structural identity must be unique. Flat headers all
    # have context=None, so their previous duplicate protection remains.
    identities = [
        ((a.structural_context or "").strip().casefold(), a.attribute_name.strip().casefold())
        for a in attributes
    ]
    duplicates = sorted({identity for identity in identities if identities.count(identity) > 1})
    if duplicates:
        labels = [f"{context} / {name}" if context else name for context, name in duplicates]
        raise ValueError(
            "Nama atribut duplikat dalam kelompok yang sama: " + ", ".join(labels)
        )


def load_transposed_rows(path: Path, sheet_name: str) -> tuple[list[ParsedAttribute], list[str]]:
    """Parse a transposed sheet (varietas as column headers, characters as
    rows): finds the header row (first row whose first cell is "Karakter"),
    then treats every subsequent row as one attribute — its varietas-column
    values are the row_values, positionally aligned with the returned
    `variety_names` (index i in both refers to the same column)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = next(
        (i for i, row in enumerate(rows) if row and str(row[0]).strip() == "Karakter"),
        None,
    )
    if header_idx is None:
        raise ValueError(f'no header row starting with "Karakter" found in {path} sheet {sheet_name!r}')

    header_row = rows[header_idx]
    variety_names = [str(v).strip() for v in header_row[1:] if v is not None]
    n_varieties = len(variety_names)

    attributes = []
    for row in rows[header_idx + 1 :]:
        if not row or row[0] is None:
            continue
        attribute_name = str(row[0]).strip()
        row_values = [
            (str(v) if v is not None else None) for v in list(row[1:])[:n_varieties]
        ]
        row_values += [None] * (n_varieties - len(row_values))  # pad if the row was short
        attributes.append(ParsedAttribute(attribute_name, None, row_values))
    return attributes, variety_names
