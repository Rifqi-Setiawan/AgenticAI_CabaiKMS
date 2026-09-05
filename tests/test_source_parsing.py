from pathlib import Path

import openpyxl
import pytest

from src.agents.schema_matching.source_parsing import load_row_oriented_columns, load_transposed_rows


@pytest.fixture
def flat_observations(tmp_path):
    """Minimal T03-shaped fixture: one header and all three observations."""
    path = tmp_path / "observations.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Observations"
    ws.append(["Sample_ID", "Variety", "Growth habit", "Plant Height (cm)"])
    ws.append(["OBS-001", "Domba", "terna", "50 - 77 cm"])
    ws.append(["OBS-002", "Gendot", "perdu", "100 cm"])
    ws.append(["OBS-003", "Kopay", "terna", None])
    wb.save(path)
    wb.close()
    return path


@pytest.mark.parametrize("header_rows", [None, 1])
def test_flat_header_preserves_first_observation(flat_observations, header_rows):
    parsed = load_row_oriented_columns(flat_observations, "Observations", header_rows=header_rows)
    attrs = {p.attribute_name: p for p in parsed}
    assert list(attrs) == ["Sample_ID", "Variety", "Growth habit", "Plant Height (cm)"]
    assert attrs["Variety"].row_values == ["Domba", "Gendot", "Kopay"]
    assert attrs["Growth habit"].row_values == ["terna", "perdu", "terna"]
    assert attrs["Plant Height (cm)"].row_values == ["50 - 77 cm", "100 cm", None]
    assert all(p.structural_context is None for p in parsed)


@pytest.mark.parametrize("merged,header_rows", [(True, None), (True, 2), (False, 2)])
def test_hierarchical_headers_still_supported(tmp_path, merged, header_rows):
    path = tmp_path / "hierarchical.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Jenis Cabai", "Morfologi", None])
    ws.append([None, "Habitus", "Tinggi"])
    ws.append(["Domba", "terna", "50 cm"])
    ws.append(["Gendot", None, "100 cm"])
    if merged:
        ws.merge_cells("A1:A2")
        ws.merge_cells("B1:C1")
    wb.save(path)
    wb.close()
    parsed = load_row_oriented_columns(path, "Sheet", header_rows=header_rows)
    assert [p.attribute_name for p in parsed] == ["Jenis Cabai", "Habitus", "Tinggi"]
    assert parsed[0].row_values == ["Domba", "Gendot"]
    assert parsed[1].row_values == ["terna", None]
    assert parsed[2].structural_context == "Morfologi"


@pytest.mark.parametrize("sample,sheet", [
    ("data_input.xlsx", "Resume Data"), ("data_input_sintetis_1.xlsx", "Sheet1"),
])
def test_existing_two_row_samples_keep_the_same_parsing(sample, sheet):
    path = Path(__file__).resolve().parents[1] / "data" / "samples" / sample
    auto = load_row_oriented_columns(path, sheet)
    explicit = load_row_oriented_columns(path, sheet, header_rows=2)
    assert auto == explicit
    assert any(a.attribute_name == "Jenis Cabai" for a in auto)
    assert all(len(a.row_values) == len(auto[0].row_values) for a in auto)


@pytest.mark.parametrize("rows,match", [
    ([], "Header kosong"),
    ([["Variety", "Height"]], "Tidak ada baris data"),
    ([["Variety", None], ["Domba", "50 cm"]], "header ambigu"),
    ([["Variety", "Variety"], ["Domba", "Gendot"]], "duplikat"),
])
def test_invalid_headers_fail_clearly(tmp_path, rows, match):
    path = tmp_path / "invalid.xlsx"
    wb = openpyxl.Workbook()
    for row in rows:
        wb.active.append(row)
    wb.save(path)
    wb.close()
    with pytest.raises(ValueError, match=match):
        load_row_oriented_columns(path, "Sheet")


def test_transposed_sample_unchanged():
    path = Path(__file__).resolve().parents[1] / "data/samples/sample_transposed_sintetis.xlsx"
    attrs, varieties = load_transposed_rows(path, "Data Transposed (Sintetis)")
    assert len(varieties) == 3
    assert attrs and all(len(a.row_values) == len(varieties) for a in attrs)


def test_invalid_header_count(flat_observations):
    with pytest.raises(ValueError, match="Jumlah baris header"):
        load_row_oriented_columns(flat_observations, "Observations", header_rows=3)


def test_repeated_leaf_headers_are_valid_under_different_sections(tmp_path):
    path = tmp_path / "multi_level_duplicates.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Identitas", None, "Young Fruit", None, "Mature Fruit", None])
    ws.append(["Variety", "Location", "Fruit Length", "Fruit Colour", "Fruit Length", "Fruit Colour"])
    ws.append(["Domba", "Jawa Barat", "3 cm", "green", "5 cm", "red"])
    ws.merge_cells("A1:B1")
    ws.merge_cells("C1:D1")
    ws.merge_cells("E1:F1")
    wb.save(path)
    wb.close()

    parsed = load_row_oriented_columns(path, "Sheet")
    assert [a.display_name for a in parsed] == [
        "Identitas / Variety", "Identitas / Location",
        "Young Fruit / Fruit Length", "Young Fruit / Fruit Colour",
        "Mature Fruit / Fruit Length", "Mature Fruit / Fruit Colour",
    ]
    young_length, mature_length = parsed[2], parsed[4]
    assert young_length.attribute_name == mature_length.attribute_name == "Fruit Length"
    assert young_length.structural_context == "Young Fruit"
    assert mature_length.structural_context == "Mature Fruit"


def test_repeated_leaf_header_in_same_section_is_rejected(tmp_path):
    path = tmp_path / "same_section_duplicate.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Identity", "Fruit", None])
    ws.append(["Variety", "Length", "Length"])
    ws.append(["Domba", "3 cm", "5 cm"])
    ws.merge_cells("B1:C1")
    wb.save(path)
    wb.close()
    with pytest.raises(ValueError, match="fruit / length"):
        load_row_oriented_columns(path, "Sheet")
