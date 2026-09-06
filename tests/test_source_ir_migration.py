import openpyxl
import pytest
from pandas.testing import assert_frame_equal

from src.agents.schema_matching.anchor import AnchorResult
from src.schema.canonical import CanonicalSchema
from src.schema.contracts import SchemaMapping
from src.schema.structure import StructureProposal
from src.ui import pipeline_runner as runner
from src.ingestion.source_migration import SourceMigrationGateError
from tests.test_source_parsing import flat_observations


def _anchor(candidates, **kwargs):
    selected = next((item.column_name for item in candidates if item.column_name == "Variety"), None)
    return AnchorResult("found" if selected else "escalate", selected, 1.0, "test")


def _isolate(monkeypatch):
    monkeypatch.setattr(runner, "detect_anchor", _anchor)
    monkeypatch.setattr(runner, "ensure_indexed", lambda *args, **kwargs: None)
    monkeypatch.setattr(runner, "retrieve", lambda *args, **kwargs: [])
    monkeypatch.setattr(runner, "build_exact_index", lambda *args, **kwargs: object())
    monkeypatch.setattr(runner, "run_pipeline", lambda *args, **kwargs: {})


def _mapping(schema, attribute, target):
    return SchemaMapping(
        source_attribute=attribute,
        source_format="row-oriented",
        target_canonical_row=target,
        confidence=0.99,
        reasoning="migration fixture",
        normalization_required=True,
    )


def _flat_proposal(data_start=2):
    return StructureProposal(
        status="RESOLVED", orientation="row-oriented",
        row_oriented={
            "table_range": "A1:D4", "header_rows": [1],
            "data_start_row": data_start, "data_end_row": 4,
            "attribute_columns": ["A", "B", "C", "D"],
            "header_bindings": [
                {"column_letter": column, "header_cells": [f"{column}1"]}
                for column in "ABCD"
            ],
        }, confidence=0.9, evidence_summary="Flat table.",
    )


def test_default_backend_remains_legacy_without_structure_call(flat_observations, monkeypatch):
    _isolate(monkeypatch)
    monkeypatch.setattr(runner, "safe_rerank", lambda *args, **kwargs: (None, {}))

    def forbidden(**kwargs):
        pytest.fail("default backend must not call structure understanding")

    result = runner.run_pipeline_ui(flat_observations, structure_llm_call=forbidden)
    assert result.source_backend == "legacy"
    assert result.source_ir_version is None
    assert result.structure_shadow is None


def test_gated_row_match_promotes_once_and_enriches_provenance(
    flat_observations, monkeypatch,
):
    _isolate(monkeypatch)
    schema = CanonicalSchema.from_template()
    height = schema.row_by_label("tinggi tanaman").id

    def mapping(profile, candidates, state, *, source_format, **kwargs):
        target = height if profile.attribute_name == "Plant Height (cm)" else "NULL"
        return _mapping(schema, profile.attribute_name, target), {}

    monkeypatch.setattr(runner, "safe_rerank", mapping)
    legacy = runner.run_pipeline_ui(flat_observations)
    calls = []

    def structure(**kwargs):
        calls.append(kwargs)
        return _flat_proposal()

    gated = runner.run_pipeline_ui(
        flat_observations,
        source_backend="source-ir-gated",
        retrieval_backend="exact",
        enable_structure_shadow=True,
        structure_llm_call=structure,
    )
    assert len(calls) == 1
    assert gated.source_backend == "source-ir-gated"
    assert gated.retrieval_backend == "exact"
    assert gated.source_ir_version == "source-ir-v1"
    assert gated.structure_shadow.status.value == "MATCH"
    assert_frame_equal(legacy.mapping_df, gated.mapping_df)
    assert_frame_equal(legacy.canonical_df, gated.canonical_df)
    assert legacy.workbook_bytes == gated.workbook_bytes
    assert all(record.source_cells == [] for record in legacy.provenance_records)
    domba = next(
        record for record in gated.provenance_records
        if record.source_attribute == "Plant Height (cm)" and record.variety == "Domba"
    )
    assert domba.source_cells == ["D2"]
    assert domba.source_attribute_id == "Observations!COL:D"
    assert domba.source_header_cells == ["D1"]
    assert domba.source_ir_version == "source-ir-v1"
    assert gated.canonical_df.loc[
        gated.canonical_df.Karakter == "tinggi tanaman", "Kopay"
    ].item() == ""


def test_gated_transposed_match_promotes_entity_alignment(tmp_path, monkeypatch):
    path = tmp_path / "transposed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T"
    ws.append(["Karakter", "Domba", "Gendot"])
    ws.append(["Habit", "terna", "perdu"])
    wb.save(path)
    wb.close()
    _isolate(monkeypatch)
    schema = CanonicalSchema.from_template()
    habitus = schema.row_by_label("habitus").id
    monkeypatch.setattr(
        runner, "safe_rerank",
        lambda profile, candidates, state, *, source_format, **kwargs: (
            _mapping(schema, profile.attribute_name, habitus), {}
        ),
    )
    proposal = StructureProposal(
        status="RESOLVED", orientation="transposed",
        transposed={
            "table_range": "A1:C2", "header_row": 1, "label_column": "A",
            "data_columns": ["B", "C"], "attribute_start_row": 2,
            "attribute_end_row": 2,
        }, confidence=0.9, evidence_summary="Transposed table.",
    )
    result = runner.run_pipeline_ui(
        path, source_format="transposed", source_backend="source-ir-gated",
        structure_llm_call=lambda **kwargs: proposal,
    )
    assert result.structure_shadow.status.value == "MATCH"
    assert result.structure_shadow.entity_names_match
    assert result.canonical_df.loc[result.canonical_df.Karakter == "habitus", "Domba"].item() == "terna"
    assert result.provenance_records[0].source_cells == ["B2"]


def test_parity_difference_fails_before_schema_matching(flat_observations, monkeypatch):
    monkeypatch.setattr(runner, "detect_anchor", _anchor)
    monkeypatch.setattr(runner, "ensure_indexed", lambda *a, **k: pytest.fail("gate must precede indexing"))
    monkeypatch.setattr(runner, "retrieve", lambda *a, **k: pytest.fail("gate must precede retrieval"))
    monkeypatch.setattr(runner, "safe_rerank", lambda *a, **k: pytest.fail("gate must precede reranking"))
    with pytest.raises(SourceMigrationGateError) as caught:
        runner.run_pipeline_ui(
            flat_observations, source_backend="source-ir-gated",
            structure_llm_call=lambda **kwargs: _flat_proposal(data_start=3),
        )
    assert caught.value.status.value == "DIFFERENT"
    assert "VALUE_POSITION_MISMATCH" in caught.value.issue_codes


@pytest.mark.parametrize("proposal", [
    StructureProposal(status="AMBIGUOUS", confidence=0.2, evidence_summary="Unclear."),
    _flat_proposal().model_copy(deep=True),
])
def test_abstention_and_verifier_failure_cannot_reach_schema_matching(
    flat_observations, monkeypatch, proposal,
):
    _isolate(monkeypatch)
    monkeypatch.setattr(runner, "ensure_indexed", lambda *a, **k: pytest.fail("must fail at gate"))
    if proposal.status.value == "RESOLVED":
        proposal.row_oriented.header_bindings[0].header_cells = ["Z999"]
    with pytest.raises(SourceMigrationGateError) as caught:
        runner.run_pipeline_ui(
            flat_observations, source_backend="source-ir-gated",
            structure_llm_call=lambda **kwargs: proposal,
        )
    assert caught.value.status.value in {"NEW_PATH_ABSTAINED", "NEW_PATH_FAILED"}


def test_legacy_failure_new_success_is_never_promoted(tmp_path, monkeypatch):
    path = tmp_path / "messy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Messy"
    ws.append(["Research title"])
    ws.append(["Location", "Bogor"])
    ws.append([])
    ws.append([None, "Identity", "Morphology"])
    ws.append([None, "Variety", "Height"])
    ws.append([None, "name", "cm"])
    ws.append([None, "Domba", 80])
    ws.append([None, "Gendot", 90])
    wb.save(path)
    wb.close()
    monkeypatch.setattr(runner, "detect_anchor", _anchor)
    monkeypatch.setattr(runner, "ensure_indexed", lambda *a, **k: pytest.fail("must not index"))
    proposal = StructureProposal(
        status="RESOLVED", orientation="row-oriented",
        row_oriented={
            "table_range": "B4:C8", "header_rows": [4, 5, 6],
            "data_start_row": 7, "data_end_row": 8,
            "attribute_columns": ["B", "C"],
            "header_bindings": [
                {"column_letter": "B", "header_cells": ["B4", "B5"]},
                {"column_letter": "C", "header_cells": ["C4", "C5"]},
            ],
        }, confidence=0.9, evidence_summary="Messy table.",
    )
    with pytest.raises(SourceMigrationGateError) as caught:
        runner.run_pipeline_ui(
            path, sheet_name="Messy", source_backend="source-ir-gated",
            structure_llm_call=lambda **kwargs: proposal,
        )
    assert caught.value.status.value == "LEGACY_FAILED"
    assert "LEGACY_REFERENCE_UNAVAILABLE" in caught.value.issue_codes
    assert caught.value.report.new_path_resolved


@pytest.mark.parametrize("values, expected_raw", [
    (("10", "12"), "10; 12"),
    (("10", "10"), "10"),
])
def test_repeated_variety_preserves_all_distinct_source_observations(
    tmp_path, monkeypatch, values, expected_raw,
):
    path = tmp_path / "repeated.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Variety", "Height"])
    ws.append(["Domba", values[0]])
    ws.append(["Domba", values[1]])
    wb.save(path)
    wb.close()
    _isolate(monkeypatch)
    schema = CanonicalSchema.from_template()
    height = schema.row_by_label("tinggi tanaman").id
    monkeypatch.setattr(
        runner, "safe_rerank",
        lambda profile, candidates, state, *, source_format, **kwargs: (
            _mapping(schema, profile.attribute_name, height), {}
        ),
    )
    proposal = StructureProposal(
        status="RESOLVED", orientation="row-oriented",
        row_oriented={
            "table_range": "A1:B3", "header_rows": [1],
            "data_start_row": 2, "data_end_row": 3,
            "attribute_columns": ["A", "B"],
            "header_bindings": [
                {"column_letter": "A", "header_cells": ["A1"]},
                {"column_letter": "B", "header_cells": ["B1"]},
            ],
        }, confidence=0.9, evidence_summary="Repeated variety.",
    )
    result = runner.run_pipeline_ui(
        path, source_backend="source-ir-gated",
        structure_llm_call=lambda **kwargs: proposal,
    )
    record = result.provenance_records[0]
    assert record.raw_value == expected_raw
    assert record.source_cells == ["B2", "B3"]
