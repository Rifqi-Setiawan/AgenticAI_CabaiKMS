from __future__ import annotations

import openpyxl
import pytest

from src.agents.tabular_update import (
    IMAGE_ROW_LABELS,
    UnknownVarietyError,
    apply_vision_result,
    apply_vision_result_to_worksheet,
    apply_vision_result_with_trace,
    apply_vision_results,
)
from src.schema.contracts import ImageMetadata, VisionResult

TEMPLATE_ROWS = [
    ["Nomor", "Karakter", "Gendot", "Kopay"],
    [1, "habitus", "perdu", "terna"],
    [2, "tinggi tanaman", "60 - 89 cm", "50 - 77 cm"],
    [56, "Lokasi", None, None],
    [57, "Gambar Daun", None, None],
    [58, "Gambar Batang", None, None],
    [59, "Gambar Buah", None, None],
    [60, "Gambar Bunga", None, None],
]


def _fresh_worksheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in TEMPLATE_ROWS:
        ws.append(row)
    return wb, ws


def _image(file_id="f1", filename="daun.jpg"):
    return ImageMetadata(
        file_id=file_id, filename=filename, mime_type="image/jpeg", size=100,
        created_time="2026-01-01T00:00:00Z",
    )


def _vision(status="KNOWN", variety="Gendot", part="DAUN", confidence=0.9, evidence="ok"):
    return VisionResult(
        classification_status=status, matched_variety=variety, identified_part=part,
        confidence=confidence, visual_evidence=evidence,
    )


def _cell(ws, label, column_letter):
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
        if row[1].value == label:
            return ws[f"{column_letter}{row[0].row}"]
    raise AssertionError(f"row {label!r} not found")


class TestKnownStatusWrites:
    def test_writes_drive_url_into_correct_cell(self):
        wb, ws = _fresh_worksheet()
        result = apply_vision_result_to_worksheet(ws, _image(file_id="abc123"), _vision())
        assert result.applied is True
        assert result.row_label == "Gambar Daun"
        assert result.column_name == "Gendot"
        cell = _cell(ws, "Gambar Daun", "C")
        assert cell.value == "https://drive.google.com/file/d/abc123/view"

    @pytest.mark.parametrize(
        "part,expected_label",
        [("DAUN", "Gambar Daun"), ("BATANG", "Gambar Batang"), ("BUAH", "Gambar Buah"), ("BUNGA", "Gambar Bunga")],
    )
    def test_every_plant_part_maps_to_correct_row(self, part, expected_label):
        wb, ws = _fresh_worksheet()
        result = apply_vision_result_to_worksheet(ws, _image(), _vision(part=part))
        assert result.row_label == expected_label
        assert IMAGE_ROW_LABELS[part] == expected_label

    def test_matches_variety_column_case_insensitively(self):
        wb, ws = _fresh_worksheet()
        result = apply_vision_result_to_worksheet(ws, _image(), _vision(variety="gendot"))
        assert result.applied is True
        assert result.column_name == "gendot"
        assert _cell(ws, "Gambar Daun", "C").value is not None


class TestNonKnownStatusNeverWrites:
    @pytest.mark.parametrize("status", ["OTHER", "UNCERTAIN"])
    def test_status_not_written_to_cell(self, status):
        wb, ws = _fresh_worksheet()
        result = apply_vision_result_to_worksheet(ws, _image(), _vision(status=status, variety=None))
        assert result.applied is False
        assert status in result.reason
        assert _cell(ws, "Gambar Daun", "C").value is None
        assert _cell(ws, "Gambar Batang", "C").value is None
        assert _cell(ws, "Gambar Buah", "C").value is None
        assert _cell(ws, "Gambar Bunga", "C").value is None

    def test_known_status_but_missing_matched_variety_is_defensive_no_op(self):
        wb, ws = _fresh_worksheet()
        result = apply_vision_result_to_worksheet(ws, _image(), _vision(status="KNOWN", variety=None))
        assert result.applied is False
        assert _cell(ws, "Gambar Daun", "C").value is None


class TestStructureAndOtherValuesUntouched:
    def test_only_the_target_cell_changes_everywhere_else_is_identical(self):
        wb, ws = _fresh_worksheet()
        before = {
            (row[0].row, cell.column): cell.value
            for row in ws.iter_rows()
            for cell in row
        }

        apply_vision_result_to_worksheet(ws, _image(), _vision(part="DAUN", variety="Gendot"))

        gambar_daun_row = next(r[0].row for r in ws.iter_rows(min_row=2, max_col=2) if r[1].value == "Gambar Daun")
        gendot_col = next(c.column for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value == "Gendot")
        changed_coord = (gambar_daun_row, gendot_col)

        after = {
            (row[0].row, cell.column): cell.value
            for row in ws.iter_rows()
            for cell in row
        }

        for coord, value_before in before.items():
            if coord == changed_coord:
                assert after[coord] != value_before  # the one cell that's allowed to change
            else:
                assert after[coord] == value_before

    def test_no_column_is_ever_added_for_unknown_variety(self):
        wb, ws = _fresh_worksheet()
        n_cols_before = ws.max_column
        result = apply_vision_result_to_worksheet(ws, _image(), _vision(variety="Varietas Baru Tak Terdaftar"))
        assert result.applied is False
        assert "not found" in result.reason
        assert ws.max_column == n_cols_before

    def test_existing_habitus_value_never_overwritten(self):
        wb, ws = _fresh_worksheet()
        apply_vision_result_to_worksheet(ws, _image(), _vision(part="DAUN", variety="Gendot"))
        assert _cell(ws, "habitus", "C").value == "perdu"
        assert _cell(ws, "tinggi tanaman", "C").value == "60 - 89 cm"


class TestMultiValueAppendAndIdempotency:
    def test_second_distinct_image_is_appended_with_separator(self):
        wb, ws = _fresh_worksheet()
        apply_vision_result_to_worksheet(ws, _image(file_id="a1"), _vision())
        apply_vision_result_to_worksheet(ws, _image(file_id="a2"), _vision())
        value = _cell(ws, "Gambar Daun", "C").value
        assert "a1" in value and "a2" in value
        assert "; " in value

    def test_rerunning_same_image_does_not_duplicate(self):
        wb, ws = _fresh_worksheet()
        apply_vision_result_to_worksheet(ws, _image(file_id="a1"), _vision())
        apply_vision_result_to_worksheet(ws, _image(file_id="a1"), _vision())
        value = _cell(ws, "Gambar Daun", "C").value
        assert value.count("a1") == 1


class TestFileBasedWrapper:
    def test_apply_vision_result_persists_to_disk(self, tmp_path):
        wb, ws = _fresh_worksheet()
        path = tmp_path / "hasil.xlsx"
        wb.save(path)

        result = apply_vision_result(path, _image(file_id="f1"), _vision())
        assert result.applied is True

        reopened = openpyxl.load_workbook(path)
        cell = _cell(reopened["Sheet1"], "Gambar Daun", "C")
        assert cell.value == "https://drive.google.com/file/d/f1/view"

    def test_apply_vision_result_does_not_save_when_not_applied(self, tmp_path):
        wb, ws = _fresh_worksheet()
        path = tmp_path / "hasil.xlsx"
        wb.save(path)
        mtime_before = path.stat().st_mtime_ns

        apply_vision_result(path, _image(), _vision(status="UNCERTAIN", variety=None))

        assert path.stat().st_mtime_ns == mtime_before

    def test_apply_vision_results_batch_saves_once(self, tmp_path):
        wb, ws = _fresh_worksheet()
        path = tmp_path / "hasil.xlsx"
        wb.save(path)

        items = [
            (_image(file_id="f1"), _vision(part="DAUN", variety="Gendot")),
            (_image(file_id="f2"), _vision(part="BUAH", variety="Kopay")),
        ]
        results = apply_vision_results(path, items)
        assert all(r.applied for r in results)

        reopened = openpyxl.load_workbook(path)["Sheet1"]
        assert _cell(reopened, "Gambar Daun", "C").value == "https://drive.google.com/file/d/f1/view"
        assert _cell(reopened, "Gambar Buah", "D").value == "https://drive.google.com/file/d/f2/view"

    def test_missing_sheet_raises(self, tmp_path):
        wb, ws = _fresh_worksheet()
        path = tmp_path / "hasil.xlsx"
        wb.save(path)
        with pytest.raises(UnknownVarietyError):
            apply_vision_result(path, _image(), _vision(), sheet_name="DoesNotExist")


class TestApplyWithTrace:
    def test_applied_result_has_empty_patch(self, tmp_path):
        wb, ws = _fresh_worksheet()
        path = tmp_path / "hasil.xlsx"
        wb.save(path)
        state = {"error_trace": []}

        result, patch = apply_vision_result_with_trace(path, _image(), _vision(), state)
        assert result.applied is True
        assert patch == {}

    def test_unapplied_result_patches_error_trace(self, tmp_path):
        wb, ws = _fresh_worksheet()
        path = tmp_path / "hasil.xlsx"
        wb.save(path)
        state = {"error_trace": ["prior"]}

        result, patch = apply_vision_result_with_trace(
            path, _image(), _vision(status="OTHER", variety=None), state
        )
        assert result.applied is False
        assert patch["error_trace"] == ["prior", result.reason]
        assert state["error_trace"] == ["prior"]  # original not mutated
