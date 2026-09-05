"""Bounded deterministic evidence views over complete workbook profiles."""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from src.ingestion.structure_geometry import (
    RangeBounds,
    StructureGeometryError,
    cell_lookup,
    content_bounds,
    parse_excel_range,
    resolve_profile_cell,
)
from src.ingestion.workbook_profiler import SheetProfile

MAX_INITIAL_CELLS = 300
MAX_ROWS_PER_REGION = 8
MAX_REQUESTED_RANGES = 3
MAX_CELLS_PER_REQUESTED_RANGE = 500
MAX_REGION_SUMMARIES = 50
MAX_MERGED_SUMMARIES = 100
MAX_STATISTICS = 50


class EvidenceRequestError(ValueError):
    pass


class EvidenceCell(BaseModel):
    coordinate: str
    source_coordinate: str
    value: Any = None
    value_type: str
    bold: bool = False
    has_fill: bool = False
    merged_range: str | None = None
    is_blank: bool = False


class RegionSummary(BaseModel):
    range: str
    rows: int
    columns: int
    non_empty_cells: int
    density: float


class CompactSheetEvidence(BaseModel):
    sheet_name: str
    content_range: str | None
    candidate_regions: list[RegionSummary]
    merged_ranges: list[dict[str, Any]]
    blank_row_runs: list[dict[str, int]]
    blank_column_runs: list[dict[str, int]]
    hidden_rows: list[int]
    hidden_columns: list[str]
    freeze_panes: str | None
    row_statistics: list[dict[str, Any]]
    column_statistics: list[dict[str, Any]]
    cells: list[EvidenceCell]
    omitted_cell_count: int
    truncated: bool
    omissions: list[str] = Field(default_factory=list)


class TargetedRangeEvidence(BaseModel):
    requested_range: str
    normalized_range: str
    cells: list[EvidenceCell]


def _evidence_cell(cell, *, coordinate: str | None = None, source_coordinate: str | None = None) -> EvidenceCell:
    return EvidenceCell(
        coordinate=coordinate or cell.coordinate,
        source_coordinate=source_coordinate or cell.coordinate,
        value=cell.value,
        value_type=cell.value_type,
        bold=cell.style.bold,
        has_fill=cell.style.fill_type is not None,
        merged_range=cell.merged_range,
        is_blank=False,
    )


def build_initial_evidence(sheet: SheetProfile) -> CompactSheetEvidence:
    """Build a bounded sheet-level view; never serialize the full profile."""
    selected_coordinates: set[str] = set()
    selected = []
    for region in sheet.candidate_regions:
        region_cells = [
            cell
            for cell in sheet.cells
            if region.min_row <= cell.row <= region.max_row
            and region.min_column <= cell.column <= region.max_column
        ]
        populated_rows = sorted({cell.row for cell in region_cells})
        early_count = (MAX_ROWS_PER_REGION + 1) // 2
        chosen_rows = populated_rows[:early_count]
        remaining = MAX_ROWS_PER_REGION - len(chosen_rows)
        if remaining:
            chosen_rows += [row for row in populated_rows[-remaining:] if row not in chosen_rows]
        for cell in region_cells:
            if cell.row not in chosen_rows or cell.coordinate in selected_coordinates:
                continue
            if len(selected) >= MAX_INITIAL_CELLS:
                break
            selected_coordinates.add(cell.coordinate)
            selected.append(cell)
        if len(selected) >= MAX_INITIAL_CELLS:
            break

    omissions: list[str] = []
    regions = sheet.candidate_regions[:MAX_REGION_SUMMARIES]
    if len(regions) < len(sheet.candidate_regions):
        omissions.append(f"{len(sheet.candidate_regions) - len(regions)} candidate region summaries omitted")
    merges = sheet.merged_ranges[:MAX_MERGED_SUMMARIES]
    if len(merges) < len(sheet.merged_ranges):
        omissions.append(f"{len(sheet.merged_ranges) - len(merges)} merged range summaries omitted")
    rows = sheet.row_profiles[:MAX_STATISTICS]
    columns = sheet.column_profiles[:MAX_STATISTICS]
    if len(rows) < len(sheet.row_profiles):
        omissions.append(f"{len(sheet.row_profiles) - len(rows)} row statistics omitted")
    if len(columns) < len(sheet.column_profiles):
        omissions.append(f"{len(sheet.column_profiles) - len(columns)} column statistics omitted")

    omitted_cells = max(0, len(sheet.cells) - len(selected))
    if omitted_cells:
        omissions.append(
            f"{omitted_cells} non-empty cells omitted; request specific bounded ranges if needed"
        )
    return CompactSheetEvidence(
        sheet_name=sheet.sheet_name,
        content_range=sheet.content_range,
        candidate_regions=[
            RegionSummary(
                range=region.range,
                rows=region.max_row - region.min_row + 1,
                columns=region.max_column - region.min_column + 1,
                non_empty_cells=region.non_empty_cells,
                density=region.density,
            )
            for region in regions
        ],
        merged_ranges=[
            {
                "range": merged.range,
                "top_left_coordinate": merged.top_left_coordinate,
                "top_left_value": merged.top_left_value,
            }
            for merged in merges
        ],
        blank_row_runs=[run.model_dump() for run in sheet.blank_row_runs],
        blank_column_runs=[run.model_dump() for run in sheet.blank_column_runs],
        hidden_rows=sheet.hidden_rows,
        hidden_columns=sheet.hidden_columns,
        freeze_panes=sheet.freeze_panes,
        row_statistics=[row.model_dump() for row in rows],
        column_statistics=[column.model_dump() for column in columns],
        cells=[_evidence_cell(cell) for cell in selected],
        omitted_cell_count=omitted_cells,
        truncated=bool(omissions),
        omissions=omissions,
    )


def validate_requested_ranges(
    sheet: SheetProfile,
    requested_ranges: list[str],
) -> list[tuple[str, RangeBounds]]:
    if not requested_ranges:
        raise EvidenceRequestError("at least one requested range is required")
    if len(requested_ranges) > MAX_REQUESTED_RANGES:
        raise EvidenceRequestError(f"at most {MAX_REQUESTED_RANGES} ranges may be requested")
    sheet_content = content_bounds(sheet)
    if sheet_content is None:
        raise EvidenceRequestError("cannot request evidence from an empty sheet")

    validated: list[tuple[str, RangeBounds]] = []
    for requested in requested_ranges:
        try:
            bounds = parse_excel_range(requested, expected_sheet=sheet.sheet_name)
        except StructureGeometryError as exc:
            raise EvidenceRequestError(str(exc)) from exc
        if bounds.cell_count > MAX_CELLS_PER_REQUESTED_RANGE:
            raise EvidenceRequestError(
                f"requested range {requested!r} contains {bounds.cell_count} cells; "
                f"limit is {MAX_CELLS_PER_REQUESTED_RANGE}"
            )
        if not bounds.intersects(sheet_content):
            raise EvidenceRequestError(f"requested range {requested!r} does not intersect sheet content")
        validated.append((requested, bounds))
    return validated


def render_targeted_evidence(
    sheet: SheetProfile,
    requested_ranges: list[str],
) -> list[TargetedRangeEvidence]:
    lookup = cell_lookup(sheet)
    output: list[TargetedRangeEvidence] = []
    for requested, bounds in validate_requested_ranges(sheet, requested_ranges):
        cells: list[EvidenceCell] = []
        for row in range(bounds.min_row, bounds.max_row + 1):
            for column in range(bounds.min_column, bounds.max_column + 1):
                coordinate = f"{get_column_letter(column)}{row}"
                resolved = resolve_profile_cell(sheet, coordinate, cells_by_coordinate=lookup)
                if resolved is None:
                    cells.append(
                        EvidenceCell(
                            coordinate=coordinate,
                            source_coordinate=coordinate,
                            value=None,
                            value_type="empty",
                            is_blank=True,
                        )
                    )
                else:
                    source_coordinate, cell = resolved
                    item = _evidence_cell(
                        cell,
                        coordinate=coordinate,
                        source_coordinate=source_coordinate,
                    )
                    cells.append(item)
        output.append(
            TargetedRangeEvidence(
                requested_range=requested,
                normalized_range=bounds.coordinate,
                cells=cells,
            )
        )
    return output
