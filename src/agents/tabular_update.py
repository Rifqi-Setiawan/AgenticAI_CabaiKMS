"""Fase 6 — Tabular Update Agent: write one Vision Agent result into the
output canonical spreadsheet's image cell. Purely deterministic, no LLM.

Operates on an already-existing OUTPUT workbook shaped like
data/canonical/template_kanonik.xlsx (same row labels, same varietas
column headers) — never on the template itself. The template is a
read-only input per CLAUDE.md's conventions; this module only ever opens
and edits the pipeline's own output file, whatever path the caller gives
it.

Each call touches AT MOST one cell: the intersection of the image row
(Gambar Daun/Batang/Buah/Bunga, chosen by VisionResult.identified_part)
and the varietas column matching VisionResult.matched_variety. Every other
cell — every other row label, every other column header, every
morphological value already filled in — is left byte-for-byte untouched.
No column is ever added: if matched_variety doesn't match an existing
column, the update is refused (see UnknownVarietyError below), never
silently turned into a new column.

Only classification_status == "KNOWN" is ever written to a cell. OTHER
(suspected new/unlisted variety) and UNCERTAIN (weak visual evidence) are
never written — writing an unconfident guess into what's supposed to be
an authoritative spreadsheet would corrupt real data. Those are instead
surfaced through GlobalState.error_trace (apply_vision_result_with_trace),
the same convention src/agents/schema_matching/review_queue.py and
normalize.py already use, rather than inventing a second review mechanism.

If the target cell already holds a value (e.g. a previous image's
reference for the same part+variety), the new reference is appended with
"; " — the same multi-value separator convention as
src/agents/schema_matching/normalize.py — rather than overwritten, and a
duplicate reference is never appended twice (idempotent re-runs).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.agents.schema_matching.review_queue import append_error_trace
from src.schema.contracts import ImageMetadata, VisionResult
from src.schema.state import GlobalState

IMAGE_ROW_LABELS: dict[str, str] = {
    "DAUN": "Gambar Daun",
    "BATANG": "Gambar Batang",
    "BUAH": "Gambar Buah",
    "BUNGA": "Gambar Bunga",
}

MULTI_VALUE_SEPARATOR = "; "


class UnknownVarietyError(Exception):
    """Raised only by the file-path convenience wrapper on an
    unrecoverable structural problem (e.g. the sheet is missing entirely);
    an unmatched variety/row is reported via TabularUpdateResult instead,
    never raised, since that's an expected, everyday outcome to route to
    review rather than a bug."""


def _image_reference_value(image: ImageMetadata) -> str:
    """What actually gets written into a Gambar cell: a Drive "view" URL
    built from file_id. Not specified by the brief — a bare file_id isn't
    directly openable and a filename alone isn't a durable reference, so
    this is a deliberate, documented choice rather than an implicit guess."""
    return f"https://drive.google.com/file/d/{image.file_id}/view"


def _find_row(ws: Worksheet, label: str) -> int | None:
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
        cell_label = row[1].value
        if cell_label is not None and str(cell_label).strip() == label:
            return row[0].row
    return None


def _find_column(ws: Worksheet, variety_name: str) -> int | None:
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    lowered = variety_name.strip().lower()
    for cell in header_row[2:]:
        if cell.value is not None and str(cell.value).strip().lower() == lowered:
            return cell.column
    return None


class TabularUpdateResult:
    def __init__(
        self,
        applied: bool,
        reason: str | None = None,
        *,
        row_label: str | None = None,
        column_name: str | None = None,
        written_value: str | None = None,
    ):
        self.applied = applied
        self.reason = reason
        self.row_label = row_label
        self.column_name = column_name
        self.written_value = written_value

    def __repr__(self) -> str:  # pragma: no cover — debugging convenience only
        return (
            f"TabularUpdateResult(applied={self.applied!r}, reason={self.reason!r}, "
            f"row_label={self.row_label!r}, column_name={self.column_name!r}, "
            f"written_value={self.written_value!r})"
        )


def apply_vision_result_to_worksheet(
    ws: Worksheet,
    image: ImageMetadata,
    vision_result: VisionResult,
) -> TabularUpdateResult:
    """Pure logic, no file I/O — operate on an already-open worksheet.
    This is what tests exercise directly; apply_vision_result() below is
    the thin file-path wrapper around it."""
    if vision_result.classification_status != "KNOWN":
        return TabularUpdateResult(
            applied=False,
            reason=(
                f"classification_status={vision_result.classification_status!r} for "
                f"file_id={image.file_id!r} — not written to spreadsheet, needs human "
                f"review (visual_evidence: {vision_result.visual_evidence})"
            ),
        )

    if vision_result.matched_variety is None:
        return TabularUpdateResult(
            applied=False,
            reason=f"KNOWN status but matched_variety is None for file_id={image.file_id!r}",
        )

    row_label = IMAGE_ROW_LABELS[vision_result.identified_part]
    row_idx = _find_row(ws, row_label)
    if row_idx is None:
        return TabularUpdateResult(applied=False, reason=f"row {row_label!r} not found in output sheet")

    col_idx = _find_column(ws, vision_result.matched_variety)
    if col_idx is None:
        return TabularUpdateResult(
            applied=False,
            reason=(
                f"varietas column {vision_result.matched_variety!r} not found — cannot "
                "add a new column (schema structure is fixed); route to review"
            ),
        )

    cell = ws.cell(row=row_idx, column=col_idx)
    new_value = _image_reference_value(image)
    existing = cell.value

    if existing is None or str(existing).strip() == "":
        cell.value = new_value
    else:
        existing_values = [v.strip() for v in str(existing).split(MULTI_VALUE_SEPARATOR)]
        if new_value not in existing_values:
            cell.value = str(existing) + MULTI_VALUE_SEPARATOR + new_value
        # else: identical reference already present — idempotent no-op

    return TabularUpdateResult(
        applied=True,
        row_label=row_label,
        column_name=vision_result.matched_variety,
        written_value=cell.value,
    )


def apply_vision_result(
    workbook_path: Path | str,
    image: ImageMetadata,
    vision_result: VisionResult,
    *,
    sheet_name: str = "Sheet1",
) -> TabularUpdateResult:
    """File-path convenience wrapper: opens `workbook_path`, applies the
    update, saves only if something was actually applied. `workbook_path`
    must point at the pipeline's OWN output file — never the read-only
    template."""
    wb = openpyxl.load_workbook(workbook_path)
    try:
        if sheet_name not in wb.sheetnames:
            raise UnknownVarietyError(f"sheet {sheet_name!r} not found in {workbook_path}")
        result = apply_vision_result_to_worksheet(wb[sheet_name], image, vision_result)
        if result.applied:
            wb.save(workbook_path)
        return result
    finally:
        wb.close()


def apply_vision_results(
    workbook_path: Path | str,
    items: list[tuple[ImageMetadata, VisionResult]],
    *,
    sheet_name: str = "Sheet1",
) -> list[TabularUpdateResult]:
    """Batch convenience: opens the workbook once, applies every item, and
    saves once at the end if anything at all was applied — avoids a
    separate disk round trip per image."""
    wb = openpyxl.load_workbook(workbook_path)
    try:
        if sheet_name not in wb.sheetnames:
            raise UnknownVarietyError(f"sheet {sheet_name!r} not found in {workbook_path}")
        ws = wb[sheet_name]
        results = [apply_vision_result_to_worksheet(ws, image, result) for image, result in items]
        if any(r.applied for r in results):
            wb.save(workbook_path)
        return results
    finally:
        wb.close()


def apply_vision_result_with_trace(
    workbook_path: Path | str,
    image: ImageMetadata,
    vision_result: VisionResult,
    state: GlobalState,
    *,
    sheet_name: str = "Sheet1",
) -> tuple[TabularUpdateResult, dict[str, Any]]:
    """Convenience for an orchestrator node: apply the update, and if it
    wasn't applied, return the GlobalState.error_trace patch recording why
    (same convention as review_queue.process_mapping / normalize_with_trace).
    The patch is {} when the update succeeded — nothing to record."""
    result = apply_vision_result(workbook_path, image, vision_result, sheet_name=sheet_name)
    patch = append_error_trace(state, result.reason) if not result.applied else {}
    return result, patch
