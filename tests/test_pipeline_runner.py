import io
import hashlib
from pathlib import Path

import openpyxl
import pytest
from pandas.testing import assert_frame_equal

from src.agents.schema_matching.anchor import AnchorResult
from src.agents.schema_matching.exact_match import ExactNameResolution, ExactNameStatus
from src.agents.schema_matching.retrieval import RetrievalHit
from src.schema.canonical import CanonicalSchema
from src.schema.contracts import SchemaMapping
from src.schema.structure import StructureProposal
from src.ui import pipeline_runner as runner
from src.ui.output_builder import worksheet_to_dataframe
from tests.test_source_parsing import flat_observations  # shared temporary workbook fixture


def _all_candidates(schema=None):
    schema = schema or CanonicalSchema.from_template()
    return [
        RetrievalHit(
            row.id,
            row.label,
            row.domain,
            index / 100.0,
            canonical_key=row.canonical_key,
        )
        for index, row in enumerate(schema.rows)
    ]


def test_flat_input_values_reach_downloaded_workbook(flat_observations, monkeypatch):
    """Run real parsing/grouping/normalization/Excel serialization, no API calls."""
    schema = CanonicalSchema.from_template()
    targets = {"Growth habit": "habitus", "Plant Height (cm)": "tinggi tanaman"}
    monkeypatch.setattr(runner, "detect_anchor", lambda *a, **kw: AnchorResult("found", "Variety", 1.0, "test"))
    monkeypatch.setattr(runner, "ensure_indexed", lambda *a, **kw: None)
    monkeypatch.setattr(runner, "retrieve", lambda *a, **kw: _all_candidates(schema))
    monkeypatch.setattr(runner, "run_pipeline", lambda *a, **kw: {})  # do not touch user checkpoints

    def mapping(profile, candidates, state, *, source_format, **kwargs):
        label = targets.get(profile.attribute_name)
        target = schema.row_by_label(label).id if label else "NULL"
        return SchemaMapping(
            source_attribute=profile.attribute_name, source_format=source_format,
            target_canonical_row=target, confidence=0.99,
            reasoning="Injected mapping: tests wiring, not model accuracy",
            normalization_required=True,
        ), {}

    monkeypatch.setattr(runner, "safe_rerank", mapping)
    result = runner.run_pipeline_ui(flat_observations)
    wb = openpyxl.load_workbook(io.BytesIO(result.workbook_bytes))
    try:
        ws = wb["Sheet1"]
        assert [ws.cell(1, c).value for c in range(3, 6)] == ["Domba", "Gendot", "Kopay"]
        row = next(c.row for c in ws["B"] if c.value == "habitus")
        assert [ws.cell(row, c).value for c in range(3, 6)] == ["terna", "perdu", "terna"]
        height_row = next(c.row for c in ws["B"] if c.value == "tinggi tanaman")
        assert ws.cell(height_row, 3).value  # first observation, previously lost
        assert ws.cell(height_row, 5).value is None
        downloaded = worksheet_to_dataframe(ws, schema, ["Domba", "Gendot", "Kopay"])
        assert_frame_equal(downloaded, result.canonical_df)
        assert "terna" not in result.mapping_df.source_attribute.tolist()
        assert result.mapping_df.source_attribute_display.tolist() == result.mapping_df.source_attribute.tolist()
        assert result.mapping_df.mapping_item_id.is_unique
        assert set(result.mapping_df.source_file_sha256) == {hashlib.sha256(flat_observations.read_bytes()).hexdigest()}
        assert set(result.mapping_df.source_sheet) == {"Observations"}
        assert all(item.mapping_item_id for item in result.mapping_verifications)
        assert len(result.provenance_records) == 5
        assert len({record.run_id for record in result.provenance_records}) == 1
        assert result.provenance_records[0].run_id == result.checkpoint_thread_id
        height_records = [
            record for record in result.provenance_records
            if record.source_attribute == "Plant Height (cm)"
        ]
        assert {record.variety for record in height_records} == {"Domba", "Gendot"}
        domba_height = next(record for record in height_records if record.variety == "Domba")
        assert domba_height.source_file_name == flat_observations.name
        assert domba_height.source_file_sha256 == hashlib.sha256(flat_observations.read_bytes()).hexdigest()
        assert domba_height.source_sheet == "Observations"
        assert domba_height.source_context is None
        assert domba_height.source_attribute_display == "Plant Height (cm)"
        assert domba_height.source_cells == []
        assert domba_height.variety == "Domba"
        assert domba_height.canonical_row_id == schema.row_by_label("tinggi tanaman").id
        assert domba_height.canonical_key == "tinggi_tanaman"
        assert domba_height.canonical_label == "tinggi tanaman"
        assert domba_height.canonical_domain == "vegetatif"
        assert domba_height.raw_value == "50 - 77 cm"
        assert domba_height.normalized_value == "50--77 cm"
        assert domba_height.normalization_required is True
        assert domba_height.mapping_confidence == 0.99
        assert domba_height.acceptance_status == "AUTO_ACCEPT"
        assert domba_height.acceptance_reason
        assert domba_height.canonical_write is True
        assert domba_height.schema_version == schema.schema_version
        assert domba_height.template_hash == schema.template_hash
        assert domba_height.mapping_method == "retrieve_rerank"
        assert set(result.mapping_df.mapping_method) == {"retrieve_rerank"}
    finally:
        wb.close()


def test_multilevel_duplicate_leaf_names_stay_distinct_through_export(tmp_path, monkeypatch):
    """A repeated leaf header is identified by its parent section, end to end."""
    source = tmp_path / "multilevel.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Identity", "Young Fruit", None, "Mature Fruit", None])
    ws.append(["Variety", "Fruit Length", "Position", "Fruit Length", "Position"])
    ws.append(["Domba", "3 cm", "pendant", "5 cm", "erect"])
    ws.merge_cells("B1:C1")
    ws.merge_cells("D1:E1")
    wb.save(source)
    wb.close()

    schema = CanonicalSchema.from_template()
    targets = {
        ("Young Fruit", "Fruit Length"): "panjang buah muda",
        ("Mature Fruit", "Fruit Length"): "panjang buah masak",
    }
    monkeypatch.setattr(runner, "detect_anchor", lambda *a, **kw: AnchorResult("found", "Variety", 1.0, "test"))
    monkeypatch.setattr(runner, "ensure_indexed", lambda *a, **kw: None)
    monkeypatch.setattr(runner, "retrieve", lambda *a, **kw: _all_candidates(schema))
    monkeypatch.setattr(runner, "run_pipeline", lambda *a, **kw: {})

    def mapping(profile, candidates, state, *, source_format, **kwargs):
        label = targets.get((profile.structural_context, profile.attribute_name))
        target = schema.row_by_label(label).id if label else "NULL"
        return SchemaMapping(
            source_attribute=profile.attribute_name,
            source_context=profile.structural_context,
            source_format=source_format,
            target_canonical_row=target,
            confidence=0.99,
            reasoning="Injected mapping: verifies structural context wiring",
            normalization_required=True,
        ), {}

    monkeypatch.setattr(runner, "safe_rerank", mapping)
    result = runner.run_pipeline_ui(source, header_rows=2)

    lengths = result.mapping_df[result.mapping_df.source_attribute == "Fruit Length"]
    assert lengths.source_attribute_display.tolist() == [
        "Young Fruit / Fruit Length",
        "Mature Fruit / Fruit Length",
    ]
    assert lengths.source_context.tolist() == ["Young Fruit", "Mature Fruit"]
    assert set(lengths.predicted_label) == {"panjang buah muda", "panjang buah masak"}

    exported = openpyxl.load_workbook(io.BytesIO(result.workbook_bytes))
    try:
        output = exported["Sheet1"]
        young_row = next(cell.row for cell in output["B"] if cell.value == "panjang buah muda")
        mature_row = next(cell.row for cell in output["B"] if cell.value == "panjang buah masak")
        assert output.cell(young_row, 3).value == "3 cm"
        assert output.cell(mature_row, 3).value == "5 cm"
    finally:
        exported.close()


def test_missing_anchor_stops_before_indexing_or_llm(flat_observations, monkeypatch):
    monkeypatch.setattr(runner, "detect_anchor", lambda *a, **kw: AnchorResult("escalate", None, 0.1, "test"))
    monkeypatch.setattr(runner, "ensure_indexed", lambda *a, **kw: pytest.fail("Must fail before indexing"))
    with pytest.raises(ValueError, match="Kolom varietas tidak ditemukan"):
        runner.run_pipeline_ui(flat_observations)


def test_missing_variety_value_stops_instead_of_losing_observation(flat_observations, monkeypatch):
    wb = openpyxl.load_workbook(flat_observations)
    wb.active["B2"] = " "
    wb.save(flat_observations)
    wb.close()
    monkeypatch.setattr(runner, "detect_anchor", lambda *a, **kw: AnchorResult("found", "Variety", 1.0, "test"))
    monkeypatch.setattr(runner, "ensure_indexed", lambda *a, **kw: pytest.fail("Must fail before indexing"))
    with pytest.raises(ValueError, match="Varietas kosong pada observasi ke-1"):
        runner.run_pipeline_ui(flat_observations)


def _isolate_pipeline(monkeypatch):
    monkeypatch.setattr(runner, "detect_anchor", lambda *a, **kw: AnchorResult("found", "Variety", 1.0, "test"))
    monkeypatch.setattr(runner, "ensure_indexed", lambda *a, **kw: None)
    monkeypatch.setattr(runner, "retrieve", lambda *a, **kw: _all_candidates())
    monkeypatch.setattr(runner, "run_pipeline", lambda *a, **kw: {})


def _mapping(schema, attribute, row_id, confidence=0.99):
    return SchemaMapping(
        source_attribute=attribute,
        source_format="row-oriented",
        target_canonical_row=row_id,
        confidence=confidence,
        reasoning="pipeline safety-gate fixture",
        normalization_required=True,
    )


def test_exact_retrieval_pipeline_is_offline_and_does_not_touch_chroma(
    flat_observations, monkeypatch,
):
    canonical_count = len(CanonicalSchema.from_template().rows)
    monkeypatch.setattr(
        runner,
        "detect_anchor",
        lambda *a, **kw: AnchorResult("found", "Variety", 1.0, "test"),
    )
    monkeypatch.setattr(
        runner,
        "ensure_indexed",
        lambda *a, **kw: pytest.fail("exact pipeline must not access Chroma"),
    )
    monkeypatch.setattr(runner, "run_pipeline", lambda *a, **kw: {})
    monkeypatch.setattr(runner, "safe_rerank", lambda *a, **kw: (None, {}))
    encode_calls = []

    def fake_encode(texts, model_name=None):
        encode_calls.append(tuple(texts))
        return [[1.0, 0.0] for _ in texts]

    result = runner.run_pipeline_ui(
        flat_observations,
        source_backend="legacy",
        retrieval_backend="exact",
        embedding_encode_call=fake_encode,
    )
    assert result.source_backend == "legacy"
    assert result.retrieval_backend == "exact"
    assert result.agent_status["retrieval"].startswith("exact — exhaustive cosine")
    assert sum(len(call) == canonical_count for call in encode_calls) == 1
    assert sum(len(call) == 1 for call in encode_calls) == 3


@pytest.mark.parametrize("source_backend", ["legacy-ish", ""])
def test_unknown_source_backend_fails_before_file_access(source_backend):
    with pytest.raises(ValueError, match="unknown source backend"):
        runner.run_pipeline_ui(Path("does-not-exist.xlsx"), source_backend=source_backend)


@pytest.mark.parametrize("retrieval_backend", ["excat", "faiss", ""])
def test_unknown_pipeline_retrieval_backend_fails_before_file_access(retrieval_backend):
    with pytest.raises(ValueError, match="unknown retrieval backend"):
        runner.run_pipeline_ui(
            Path("does-not-exist.xlsx"), retrieval_backend=retrieval_backend
        )


def test_review_mapping_never_normalizes_or_writes_but_later_attributes_do(
    flat_observations, monkeypatch,
):
    """Mixed workbook regression: ACCEPT, REVIEW, ACCEPT stay independent."""
    _isolate_pipeline(monkeypatch)
    schema = CanonicalSchema.from_template()
    accepted_rows = {
        "Sample_ID": schema.row_by_label("Lokasi").id,
        "Plant Height (cm)": schema.row_by_label("tinggi tanaman").id,
    }
    review_row = schema.row_by_label("habitus").id

    def safe_mapping(profile, candidates, state, *, source_format, **kwargs):
        if profile.attribute_name == "Growth habit":
            mapping = _mapping(schema, profile.attribute_name, review_row, confidence=0.1)
            return mapping, {"error_trace": ["confidence 0.10 di bawah ambang 0.6"]}
        return _mapping(schema, profile.attribute_name, accepted_rows[profile.attribute_name]), {}

    real_normalize = runner.normalize
    normalized_raw_values = []

    def normalize_spy(raw_value, target_row):
        normalized_raw_values.append(raw_value)
        return real_normalize(raw_value, target_row)

    monkeypatch.setattr(runner, "safe_rerank", safe_mapping)
    monkeypatch.setattr(runner, "normalize", normalize_spy)
    result = runner.run_pipeline_ui(flat_observations)

    review = result.mapping_df[result.mapping_df.source_attribute == "Growth habit"].iloc[0]
    later = result.mapping_df[result.mapping_df.source_attribute == "Plant Height (cm)"].iloc[0]
    assert review.acceptance_status == "REVIEW"
    assert review.canonical_write == False  # noqa: E712 - numpy bool comparison is intentional
    assert review.mapping_method == "retrieve_rerank"
    assert "confidence" in review.acceptance_reason
    assert not any("terna" in str(value) or "perdu" in str(value) for value in normalized_raw_values)
    assert result.canonical_df.loc[result.canonical_df.Karakter == "habitus", "Domba"].item() == ""
    assert not any(record.source_attribute == "Growth habit" for record in result.provenance_records)
    assert later.acceptance_status == "AUTO_ACCEPT"
    assert later.canonical_write == True  # noqa: E712
    assert result.canonical_df.loc[result.canonical_df.Karakter == "tinggi tanaman", "Domba"].item() == "50--77 cm"
    assert "2 AUTO_ACCEPT, 1 REVIEW, 0 NO_WRITE" in result.agent_status["schema_matching"]


def test_structurally_invalid_result_is_observable_and_does_not_block_next_attribute(
    flat_observations, monkeypatch,
):
    _isolate_pipeline(monkeypatch)
    schema = CanonicalSchema.from_template()
    habitus_id = schema.row_by_label("habitus").id
    height_id = schema.row_by_label("tinggi tanaman").id

    def safe_mapping(profile, candidates, state, *, source_format, **kwargs):
        if profile.attribute_name == "Sample_ID":
            return None, {"error_trace": ["format invalid; diarahkan ke manual_review"]}
        target = habitus_id if profile.attribute_name == "Growth habit" else height_id
        return _mapping(schema, profile.attribute_name, target), {}

    monkeypatch.setattr(runner, "safe_rerank", safe_mapping)
    result = runner.run_pipeline_ui(flat_observations)

    invalid = result.mapping_df[result.mapping_df.source_attribute == "Sample_ID"].iloc[0]
    following = result.mapping_df[result.mapping_df.source_attribute == "Growth habit"].iloc[0]
    assert invalid.acceptance_status == "NO_WRITE"
    assert invalid.canonical_write == False  # noqa: E712
    assert "MAPPING_MISSING" in invalid.acceptance_reason
    assert "RERANK_RELIABILITY_PATCH" in invalid.verifier_warnings
    assert not any(record.source_attribute == "Sample_ID" for record in result.provenance_records)
    assert following.acceptance_status == "AUTO_ACCEPT"
    assert following.canonical_write == True  # noqa: E712
    assert result.canonical_df.loc[result.canonical_df.Karakter == "habitus", "Domba"].item() == "terna"
    assert "2 AUTO_ACCEPT, 0 REVIEW, 1 NO_WRITE" in result.agent_status["schema_matching"]


def test_exact_alias_shortcut_still_writes_without_review(tmp_path, monkeypatch):
    source = tmp_path / "exact-alias.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Variety", "Seeds per mature fruit"])
    ws.append(["Domba", "42"])
    wb.save(source)
    wb.close()

    _isolate_pipeline(monkeypatch)
    result = runner.run_pipeline_ui(source)
    second_result = runner.run_pipeline_ui(source)

    row = result.mapping_df.iloc[0]
    assert row.acceptance_status == "AUTO_ACCEPT"
    assert row.canonical_write == True  # noqa: E712
    assert row.confidence == 1.0
    assert result.canonical_df.loc[
        result.canonical_df.Karakter == "jumlah biji/buah masak", "Domba"
    ].item() == "42"
    assert len({record.run_id for record in result.provenance_records}) == 1
    assert result.provenance_records[0].run_id != second_result.provenance_records[0].run_id
    assert row.mapping_method == "exact_name"
    assert result.provenance_records[0].mapping_method == "exact_name"
    assert result.provenance_records[0].verifier_status == "PASS"
    assert result.provenance_records[0].verifier_hard_issues == []
    assert len(result.mapping_verifications) == 1
    assert result.mapping_verifications[0].status.value == "PASS"
    assert row.mapping_item_id == result.mapping_verifications[0].mapping_item_id
    assert row.mapping_item_id == second_result.mapping_df.iloc[0].mapping_item_id
    assert result.mapping_verifications[0].retrieval_evidence is None
    assert result.agent_status["retrieval"].startswith("chroma — tidak diinisialisasi")


@pytest.mark.parametrize("retrieval_backend", ["chroma", "exact"])
def test_all_exact_name_workbook_never_initializes_retrieval_or_llm(
    tmp_path, monkeypatch, retrieval_backend,
):
    source = tmp_path / f"all-exact-{retrieval_backend}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Variety", "Seeds per mature fruit"])
    ws.append(["Domba", "42"])
    wb.save(source)
    wb.close()

    monkeypatch.setattr(
        runner, "detect_anchor",
        lambda *a, **kw: AnchorResult("found", "Variety", 1.0, "test"),
    )
    monkeypatch.setattr(runner, "run_pipeline", lambda *a, **kw: {})

    def forbidden(*args, **kwargs):
        pytest.fail("all-exact workbook must skip retrieval initialization and reranking")

    monkeypatch.setattr(runner, "ensure_indexed", forbidden)
    monkeypatch.setattr(runner, "build_exact_index", forbidden)
    monkeypatch.setattr(runner, "retrieve", forbidden)
    monkeypatch.setattr(runner, "safe_rerank", forbidden)
    result = runner.run_pipeline_ui(source, retrieval_backend=retrieval_backend)
    assert result.mapping_df.iloc[0].mapping_method == "exact_name"
    assert result.provenance_records[0].mapping_method == "exact_name"


def test_first_unresolved_attribute_initializes_chroma_once(tmp_path, monkeypatch):
    source = tmp_path / "lazy-retrieval.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Variety", "Seeds per mature fruit", "Unknown B", "Unknown C"])
    ws.append(["Domba", "42", "x", "y"])
    wb.save(source)
    wb.close()
    monkeypatch.setattr(
        runner, "detect_anchor",
        lambda *a, **kw: AnchorResult("found", "Variety", 1.0, "test"),
    )
    monkeypatch.setattr(runner, "run_pipeline", lambda *a, **kw: {})
    prepared = object()
    ensure_calls = []
    retrieve_calls = []
    rerank_calls = []

    def ensure(*args, **kwargs):
        ensure_calls.append(1)
        return prepared

    def retrieve(profile, **kwargs):
        retrieve_calls.append((profile.attribute_name, kwargs.get("collection")))
        return []

    def rerank(profile, *args, **kwargs):
        rerank_calls.append(profile.attribute_name)
        return None, {}

    monkeypatch.setattr(runner, "ensure_indexed", ensure)
    monkeypatch.setattr(runner, "retrieve", retrieve)
    monkeypatch.setattr(runner, "safe_rerank", rerank)
    result = runner.run_pipeline_ui(source)
    assert ensure_calls == [1]
    assert retrieve_calls == [("Unknown B", prepared), ("Unknown C", prepared)]
    assert rerank_calls == ["Unknown B", "Unknown C"]
    assert result.mapping_df.mapping_method.tolist() == [
        "exact_name", "retrieve_rerank", "retrieve_rerank"
    ]


def test_ambiguous_exact_name_is_observable_and_falls_through(tmp_path, monkeypatch):
    source = tmp_path / "ambiguous-name.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Variety", "Length"])
    ws.append(["Domba", "10"])
    wb.save(source)
    wb.close()
    monkeypatch.setattr(
        runner, "detect_anchor",
        lambda *a, **kw: AnchorResult("found", "Variety", 1.0, "test"),
    )
    monkeypatch.setattr(runner, "run_pipeline", lambda *a, **kw: {})
    monkeypatch.setattr(runner, "ensure_indexed", lambda *a, **kw: object())
    calls = []
    monkeypatch.setattr(
        runner,
        "resolve_exact_name",
        lambda *a, **kw: ExactNameResolution(
            status=ExactNameStatus.AMBIGUOUS,
            normalized_source_name="length",
            candidate_row_ids=("r_8", "r_31"),
            candidate_canonical_keys=("panjang_daun", "panjang_buah_masak"),
        ),
    )
    monkeypatch.setattr(
        runner, "retrieve", lambda *a, **kw: calls.append("retrieve") or []
    )
    monkeypatch.setattr(
        runner, "safe_rerank", lambda *a, **kw: calls.append("rerank") or (None, {})
    )
    result = runner.run_pipeline_ui(source)
    assert calls == ["retrieve", "rerank"]
    row = result.mapping_df.iloc[0]
    assert row.exact_name_status == "AMBIGUOUS"
    assert row.exact_name_candidates == ["panjang_daun", "panjang_buah_masak"]
    assert row.mapping_method == "retrieve_rerank"


def test_out_of_candidate_target_is_hard_blocked_before_normalization(
    tmp_path, monkeypatch,
):
    source = tmp_path / "out-of-candidate.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Variety", "Unknown attribute"])
    ws.append(["Domba", "value"])
    wb.save(source)
    wb.close()
    schema = CanonicalSchema.from_template()
    candidates = _all_candidates(schema)[:3]
    monkeypatch.setattr(
        runner, "detect_anchor",
        lambda *a, **kw: AnchorResult("found", "Variety", 1.0, "test"),
    )
    monkeypatch.setattr(runner, "ensure_indexed", lambda *a, **kw: object())
    monkeypatch.setattr(runner, "retrieve", lambda *a, **kw: candidates)
    monkeypatch.setattr(runner, "run_pipeline", lambda *a, **kw: {})
    monkeypatch.setattr(
        runner,
        "safe_rerank",
        lambda profile, *a, **kw: (
            _mapping(schema, profile.attribute_name, "r_40", confidence=0.99),
            {},
        ),
    )

    def forbidden(*args, **kwargs):
        pytest.fail("hard verifier rejection must block canonical mutation")

    monkeypatch.setattr(runner, "normalize", forbidden)
    monkeypatch.setattr(runner.CanonicalOutputBuilder, "set_cell", forbidden)
    result = runner.run_pipeline_ui(source)
    row = result.mapping_df.iloc[0]
    assert row.verifier_status == "REJECT"
    assert row.verifier_hard_issues == ["TARGET_NOT_IN_RETRIEVED_CANDIDATES"]
    assert row.acceptance_status == "NO_WRITE"
    assert row.canonical_write == False  # noqa: E712
    assert result.provenance_records == []
    assert result.mapping_verifications[0].retrieval_evidence.target_in_candidates is False
    assert "hard-blocked=1" in result.agent_status["mapping_verifier"]


def test_duplicate_noop_does_not_mark_write_or_create_provenance(tmp_path, monkeypatch):
    source = tmp_path / "duplicate-noop.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Variety", "First habit", "Duplicate habit"])
    ws.append(["Domba", "perdu", "perdu"])
    wb.save(source)
    wb.close()

    _isolate_pipeline(monkeypatch)
    schema = CanonicalSchema.from_template()
    target = schema.row_by_label("habitus").id
    monkeypatch.setattr(
        runner,
        "safe_rerank",
        lambda profile, candidates, state, *, source_format, **kwargs: (
            _mapping(schema, profile.attribute_name, target),
            {},
        ),
    )

    result = runner.run_pipeline_ui(source)
    first = result.mapping_df[result.mapping_df.source_attribute == "First habit"].iloc[0]
    duplicate = result.mapping_df[result.mapping_df.source_attribute == "Duplicate habit"].iloc[0]

    assert first.canonical_write == True  # noqa: E712
    assert duplicate.canonical_write == False  # noqa: E712
    assert len(result.provenance_records) == 1
    assert result.provenance_records[0].source_attribute == "First habit"


def test_acceptance_accounting_separates_review_and_no_write(tmp_path, monkeypatch):
    source = tmp_path / "all-acceptance-statuses.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Variety", "Accepted", "Needs review", "Invalid"])
    ws.append(["Domba", "perdu", "100 cm", "unknown"])
    wb.save(source)
    wb.close()

    _isolate_pipeline(monkeypatch)
    schema = CanonicalSchema.from_template()
    habitus = schema.row_by_label("habitus").id
    height = schema.row_by_label("tinggi tanaman").id

    def safe_mapping(profile, candidates, state, *, source_format, **kwargs):
        if profile.attribute_name == "Accepted":
            return _mapping(schema, profile.attribute_name, habitus), {}
        if profile.attribute_name == "Needs review":
            return _mapping(schema, profile.attribute_name, height, confidence=0.1), {
                "error_trace": ["confidence below threshold"]
            }
        return None, {"error_trace": ["invalid structured mapping; manual_review"]}

    monkeypatch.setattr(runner, "safe_rerank", safe_mapping)
    result = runner.run_pipeline_ui(source)

    assert result.agent_status["schema_matching"] == (
        "selesai — 3 atribut: 1 AUTO_ACCEPT, 1 REVIEW, 1 NO_WRITE; methods: "
        "0 exact_name, 3 retrieve_rerank"
    )
    assert len(result.provenance_records) == 1
    assert result.provenance_records[0].source_attribute == "Accepted"


def _shadow_flat_proposal():
    return StructureProposal(
        status="RESOLVED", orientation="row-oriented",
        row_oriented={
            "table_range": "A1:D4", "header_rows": [1],
            "data_start_row": 2, "data_end_row": 4,
            "attribute_columns": ["A", "B", "C", "D"],
            "header_bindings": [
                {"column_letter": column, "header_cells": [f"{column}1"]}
                for column in "ABCD"
            ],
        }, confidence=0.9, evidence_summary="Flat observation table.",
    )


def test_structure_shadow_is_disabled_by_default(flat_observations, monkeypatch):
    _isolate_pipeline(monkeypatch)
    monkeypatch.setattr(runner, "safe_rerank", lambda *args, **kwargs: (None, {}))
    monkeypatch.setattr(
        runner,
        "run_structure_shadow",
        lambda *args, **kwargs: pytest.fail("shadow must not run by default"),
    )
    result = runner.run_pipeline_ui(flat_observations)
    assert result.structure_shadow is None
    assert "structure_shadow" not in result.agent_status


def test_shadow_mode_cannot_change_primary_pipeline_outputs(flat_observations, monkeypatch):
    _isolate_pipeline(monkeypatch)
    schema = CanonicalSchema.from_template()
    height = schema.row_by_label("tinggi tanaman").id

    def mapping(profile, candidates, state, *, source_format, **kwargs):
        target = height if profile.attribute_name == "Plant Height (cm)" else "NULL"
        return _mapping(schema, profile.attribute_name, target), {}

    monkeypatch.setattr(runner, "safe_rerank", mapping)
    baseline = runner.run_pipeline_ui(flat_observations)
    shadowed = runner.run_pipeline_ui(
        flat_observations,
        enable_structure_shadow=True,
        structure_llm_call=lambda **kwargs: _shadow_flat_proposal(),
    )
    assert shadowed.structure_shadow is not None
    assert shadowed.structure_shadow.status.value == "MATCH"
    assert baseline.workbook_bytes == shadowed.workbook_bytes
    assert_frame_equal(baseline.canonical_df, shadowed.canonical_df)
    assert_frame_equal(baseline.mapping_df, shadowed.mapping_df)
    assert baseline.vision_rows == shadowed.vision_rows
    assert baseline.agent_status["vision_classification"] == shadowed.agent_status["vision_classification"]
    assert len(baseline.provenance_records) == len(shadowed.provenance_records)
    for first, second in zip(baseline.provenance_records, shadowed.provenance_records):
        assert first.model_dump(exclude={"run_id"}) == second.model_dump(exclude={"run_id"})


def test_shadow_exception_cannot_abort_primary_pipeline(flat_observations, monkeypatch):
    _isolate_pipeline(monkeypatch)
    monkeypatch.setattr(runner, "safe_rerank", lambda *args, **kwargs: (None, {}))

    def fail(**kwargs):
        raise RuntimeError("token=supersecret injected shadow failure")

    result = runner.run_pipeline_ui(
        flat_observations,
        enable_structure_shadow=True,
        structure_llm_call=fail,
    )
    assert result.canonical_df is not None
    assert result.workbook_bytes
    assert result.structure_shadow is not None
    assert result.structure_shadow.status.value == "NEW_PATH_FAILED"
    assert result.structure_shadow.new_path_error_type == "RuntimeError"
    assert "supersecret" not in result.structure_shadow.new_path_error_message
