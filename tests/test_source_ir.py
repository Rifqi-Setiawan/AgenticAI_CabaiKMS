import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from src.agents.schema_matching.source_parsing import (
    load_row_oriented_columns,
    load_transposed_rows,
)
from src.ingestion.source_ir_builder import build_source_ir
from src.ingestion.structure_geometry import cell_lookup, resolve_profile_cell
from src.ingestion.structure_verifier import verify_structure
from src.ingestion.workbook_profiler import profile_workbook
from src.schema.structure import StructureProposal, VerifiedStructure


def _verified(profile, proposal):
    verification = verify_structure(profile.sheets[0], proposal)
    assert verification.valid, verification
    return VerifiedStructure(proposal=proposal, verification=verification)


def test_messy_row_ir_preserves_hierarchy_repeated_leaves_and_blanks(tmp_path):
    path = tmp_path / "messy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Messy"
    ws.append(["Research report"])
    ws.append(["Location", "Bogor"])
    ws.append([])
    ws.append([None, "Identity", "Young Fruit", None, "Mature Fruit", None])
    ws.append([None, "Variety", "Size", None, "Size", None])
    ws.append([None, "Name", "Length", "Width", "Length", "Width"])
    ws.append([None, "Domba", 3, 2, 5, 4])
    ws.append([None, "Gendot", None, 3, 6, None])
    ws.append([None, "Kopay", 4, 2, 7, 5])
    ws.merge_cells("B4:B5")
    ws.merge_cells("C4:D4")
    ws.merge_cells("E4:F4")
    ws.merge_cells("C5:D5")
    ws.merge_cells("E5:F5")
    wb.save(path)
    wb.close()
    profile = profile_workbook(path)
    proposal = StructureProposal(
        status="RESOLVED",
        orientation="row-oriented",
        row_oriented={
            "table_range": "B4:F9",
            "header_rows": [4, 5, 6],
            "data_start_row": 7,
            "data_end_row": 9,
            "attribute_columns": ["B", "C", "D", "E", "F"],
            "header_bindings": [
                {"column_letter": "B", "header_cells": ["B4", "B6"]},
                {"column_letter": "C", "header_cells": ["C4", "C5", "C6"]},
                {"column_letter": "D", "header_cells": ["C4", "C5", "D6"]},
                {"column_letter": "E", "header_cells": ["E4", "E5", "E6"]},
                {"column_letter": "F", "header_cells": ["E4", "E5", "F6"]},
            ],
        },
        confidence=0.91,
        reason_codes=["MULTILEVEL_HEADER"],
        evidence_summary="Three header levels.",
    )
    verified = _verified(profile, proposal)
    ir = build_source_ir(profile, "Messy", verified)
    table = ir.tables[0]
    assert [item.row for item in table.observation_positions] == [7, 8, 9]
    young, mature = table.attributes[1], table.attributes[3]
    assert young.raw_label == mature.raw_label == "Length"
    assert young.header_path == ["Young Fruit", "Size", "Length"]
    assert mature.header_path == ["Mature Fruit", "Size", "Length"]
    assert young.source_attribute_id == "Messy!COL:C"
    assert mature.source_attribute_id == "Messy!COL:E"
    assert young.header_cells == ["C4", "C5", "C6"]
    assert [(v.coordinate, v.source_coordinate, v.raw_value, v.value_type) for v in young.values] == [
        ("C7", "C7", 3, "integer"),
        ("C8", None, None, "empty"),
        ("C9", "C9", 4, "integer"),
    ]
    assert build_source_ir(profile, "Messy", verified).model_dump(mode="json") == ir.model_dump(mode="json")


def test_transposed_ir_preserves_entity_and_value_coordinates(tmp_path):
    path = tmp_path / "transposed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T"
    ws.append(["Title"])
    ws.append([])
    ws.append([None, "Descriptor", "Domba", "Gendot", "Kopay"])
    ws.append([None, "Habit", "terna", "perdu", "terna"])
    ws.append([None, "Height", 80, None, 95])
    wb.save(path)
    wb.close()
    profile = profile_workbook(path)
    proposal = StructureProposal(
        status="RESOLVED",
        orientation="transposed",
        transposed={
            "table_range": "B3:E5", "header_row": 3, "label_column": "B",
            "data_columns": ["C", "D", "E"], "attribute_start_row": 4, "attribute_end_row": 5,
        },
        confidence=0.85,
        evidence_summary="Entities are columns.",
    )
    ir = build_source_ir(profile, "T", _verified(profile, proposal))
    table = ir.tables[0]
    assert [(p.header_coordinate, p.raw_entity_label) for p in table.observation_positions] == [
        ("C3", "Domba"), ("D3", "Gendot"), ("E3", "Kopay")
    ]
    height = table.attributes[1]
    assert height.source_attribute_id == "T!ROW:5"
    assert height.header_cells == ["B5"]
    assert [(v.coordinate, v.source_coordinate, v.raw_value, v.value_type) for v in height.values] == [
        ("C5", "C5", 80, "integer"),
        ("D5", None, None, "empty"),
        ("E5", "E5", 95, "integer"),
    ]


def test_merged_data_value_preserves_logical_and_physical_coordinates(tmp_path):
    path = tmp_path / "merged-value.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged"
    ws["D6"] = "Measurement"
    ws["C7"] = 12
    ws.merge_cells("C7:D7")
    ws["C8"] = "row-boundary"
    wb.save(path)
    wb.close()
    profile = profile_workbook(path)
    proposal = StructureProposal(
        status="RESOLVED",
        orientation="row-oriented",
        row_oriented={
            "table_range": "D6:D8",
            "header_rows": [6],
            "data_start_row": 7,
            "data_end_row": 8,
            "attribute_columns": ["D"],
            "header_bindings": [
                {"column_letter": "D", "header_cells": ["D6"]}
            ],
        },
        confidence=0.9,
        evidence_summary="One measurement column.",
    )
    values = build_source_ir(
        profile, "Merged", _verified(profile, proposal)
    ).tables[0].attributes[0].values
    assert values[0].coordinate == "D7"
    assert values[0].source_coordinate == "C7"
    assert values[0].raw_value == 12
    assert values[1].coordinate == "D8"
    assert values[1].source_coordinate is None
    assert values[1].raw_value is None
    assert values[1].value_type == "empty"


def test_unverified_structure_cannot_be_used(tmp_path):
    with pytest.raises(ValueError, match="VerifiedStructure"):
        VerifiedStructure(
            proposal=StructureProposal(
                status="AMBIGUOUS", confidence=0.1, evidence_summary="Unclear."
            ),
            verification={"valid": False},
        )


def test_legacy_flat_and_transposed_parity(tmp_path):
    flat = tmp_path / "flat.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flat"
    ws.append(["Variety", "Height"])
    ws.append(["Domba", 80])
    ws.append(["Gendot", None])
    wb.save(flat)
    wb.close()
    profile = profile_workbook(flat)
    proposal = StructureProposal(
        status="RESOLVED", orientation="row-oriented",
        row_oriented={
            "table_range": "A1:B3", "header_rows": [1], "data_start_row": 2, "data_end_row": 3,
            "attribute_columns": ["A", "B"],
            "header_bindings": [
                {"column_letter": "A", "header_cells": ["A1"]},
                {"column_letter": "B", "header_cells": ["B1"]},
            ],
        },
        confidence=1, evidence_summary="Flat table.",
    )
    attrs = build_source_ir(profile, "Flat", _verified(profile, proposal)).tables[0].attributes
    legacy = load_row_oriented_columns(flat, "Flat")
    assert [a.raw_label for a in attrs] == [a.attribute_name for a in legacy]
    assert [[None if v.raw_value is None else str(v.raw_value) for v in a.values] for a in attrs] == [
        a.row_values for a in legacy
    ]

    transposed = tmp_path / "legacy-t.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T"
    ws.append(["Karakter", "Domba", "Gendot"])
    ws.append(["Habit", "terna", None])
    wb.save(transposed)
    wb.close()
    profile = profile_workbook(transposed)
    proposal = StructureProposal(
        status="RESOLVED", orientation="transposed",
        transposed={
            "table_range": "A1:C2", "header_row": 1, "label_column": "A",
            "data_columns": ["B", "C"], "attribute_start_row": 2, "attribute_end_row": 2,
        }, confidence=1, evidence_summary="Transposed table.",
    )
    attrs = build_source_ir(profile, "T", _verified(profile, proposal)).tables[0].attributes
    legacy, names = load_transposed_rows(transposed, "T")
    assert [p.raw_entity_label for p in build_source_ir(profile, "T", _verified(profile, proposal)).tables[0].observation_positions] == names
    assert [a.raw_label for a in attrs] == [a.attribute_name for a in legacy]
    assert [[None if v.raw_value is None else str(v.raw_value) for v in a.values] for a in attrs] == [a.row_values for a in legacy]


def test_existing_two_row_sample_parity():
    from pathlib import Path

    path = Path(__file__).resolve().parents[1] / "data/samples/data_input.xlsx"
    profile = profile_workbook(path)
    sheet = next(item for item in profile.sheets if item.sheet_name == "Resume Data")
    lookup = cell_lookup(sheet)
    columns = []
    bindings = []
    current_parent = None
    for column_index in range(sheet.content_min_column, sheet.content_max_column + 1):
        column = get_column_letter(column_index)
        row_one = resolve_profile_cell(sheet, f"{column}1", cells_by_coordinate=lookup)
        row_two_cell = lookup.get(f"{column}2")
        row_two = (row_two_cell.coordinate, row_two_cell) if row_two_cell is not None else None
        if row_one is not None:
            current_parent = row_one[0]
        if row_two is not None:
            header_cells = ([current_parent] if current_parent else []) + [row_two[0]]
        elif lookup.get(f"{column}1") is not None:
            header_cells = [row_one[0]]
        else:
            continue
        columns.append(column)
        bindings.append({"column_letter": column, "header_cells": header_cells})
    proposal = StructureProposal(
        status="RESOLVED", orientation="row-oriented",
        row_oriented={
            "table_range": sheet.content_range,
            "header_rows": [1, 2], "data_start_row": 3,
            "data_end_row": sheet.content_max_row,
            "attribute_columns": columns, "header_bindings": bindings,
        },
        confidence=1,
        evidence_summary="Manually supplied legacy-compatible two-row structure.",
    )
    source_attrs = build_source_ir(
        profile, "Resume Data", _verified(profile, proposal)
    ).tables[0].attributes
    legacy = load_row_oriented_columns(path, "Resume Data", header_rows=2)
    assert [(item.raw_label, item.structural_context) for item in source_attrs] == [
        (item.attribute_name, item.structural_context) for item in legacy
    ]
    source_values = [
        [None if value.raw_value is None else str(value.raw_value) for value in item.values]
        for item in source_attrs
    ]
    for source, old in zip(source_values, legacy):
        assert old.row_values[:len(source)] == source
        assert all(value is None for value in old.row_values[len(source):])
