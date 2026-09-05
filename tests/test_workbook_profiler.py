from __future__ import annotations

from datetime import date, datetime, time

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.ingestion.workbook_profiler import (
    PROFILE_VERSION,
    WorkbookProfilerCompatibilityError,
    _iter_instantiated_cells,
    profile_sheet,
    profile_workbook,
)
from src.schema.provenance import source_file_sha256


def _save(workbook, path):
    workbook.save(path)
    workbook.close()
    return path


def _cell(sheet, coordinate):
    return next(cell for cell in sheet.cells if cell.coordinate == coordinate)


def test_normal_flat_table_has_bounds_cells_statistics_and_one_region(tmp_path):
    path = tmp_path / "flat.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Observations"
    sheet.append(["Variety", "Count", "Height", "Selected"])
    sheet.append(["Domba", 1, 50.5, True])
    sheet.append(["Gendot", 2, 70.0, False])
    sheet.append(["Kopay", 3, 80.25, True])
    _save(workbook, path)

    profile = profile_workbook(path)
    result = profile.sheets[0]

    assert profile.profile_version == PROFILE_VERSION
    assert profile.source_file_name == "flat.xlsx"
    assert result.content_range == "A1:D4"
    assert result.non_empty_cell_count == 16
    assert [cell.coordinate for cell in result.cells[:4]] == ["A1", "B1", "C1", "D1"]
    assert result.row_profiles[1].model_dump() == {
        "row_index": 2,
        "non_empty_count": 4,
        "text_count": 1,
        "numeric_count": 2,
        "boolean_count": 1,
        "formula_count": 0,
        "first_non_empty_column": 1,
        "last_non_empty_column": 4,
        "is_blank": False,
    }
    assert result.column_profiles[0].non_empty_count == 4
    assert result.column_profiles[0].text_count == 4
    assert result.candidate_regions[0].range == "A1:D4"
    assert result.candidate_regions[0].density == 1.0
    assert result.merged_ranges == []
    assert result.hidden_rows == []
    assert result.hidden_columns == []


def test_title_before_offset_table_preserves_merge_blank_gap_and_regions(tmp_path):
    path = tmp_path / "title-table.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Chili Research Data"
    sheet.merge_cells("A1:D1")
    sheet["A2"] = "Researcher: Example"
    for column, value in enumerate(["Variety", "Height", "Colour", "Count"], start=2):
        sheet.cell(4, column, value)
    for row in range(5, 9):
        for column in range(2, 6):
            sheet.cell(row, column, f"r{row}c{column}")
    _save(workbook, path)

    result = profile_workbook(path).sheets[0]

    assert result.content_range == "A1:E8"
    assert result.merged_ranges[0].range == "A1:D1"
    assert result.merged_ranges[0].top_left_value == "Chili Research Data"
    assert result.blank_row_runs[0].model_dump() == {"start": 3, "end": 3}
    assert [region.range for region in result.candidate_regions] == ["A1:A2", "B4:E8"]
    assert result.content_min_column == 1
    assert next(region for region in result.candidate_regions if region.min_row == 4).min_column == 2


def test_arbitrary_table_start_is_reported_exactly(tmp_path):
    path = tmp_path / "offset.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["D7"] = "Header"
    sheet["E7"] = "Value"
    sheet["D8"] = "Domba"
    sheet["E8"] = 42
    _save(workbook, path)

    result = profile_workbook(path).sheets[0]
    assert result.content_min_row == 7
    assert result.content_min_column == 4
    assert result.content_range == "D7:E8"


def test_multilevel_merged_headers_capture_only_physical_facts(tmp_path):
    path = tmp_path / "merged-headers.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["B2"] = "Morphology"
    sheet.merge_cells("B2:F2")
    sheet["B3"] = "Young Fruit"
    sheet.merge_cells("B3:C3")
    sheet["D3"] = "Mature Fruit"
    sheet.merge_cells("D3:E3")
    sheet["F3"] = "Other"
    for column, value in enumerate(["Length", "Colour", "Length", "Colour", "Position"], start=2):
        sheet.cell(4, column, value)
    _save(workbook, path)

    result = profile_workbook(path).sheets[0]
    assert [merge.range for merge in result.merged_ranges] == ["B2:F2", "B3:C3", "D3:E3"]
    assert [merge.top_left_value for merge in result.merged_ranges] == [
        "Morphology", "Young Fruit", "Mature Fruit"
    ]
    assert _cell(result, "B2").merged_range == "B2:F2"
    assert _cell(result, "D3").value == "Mature Fruit"
    assert not hasattr(result, "headers")
    assert not hasattr(result, "tables")


def test_blank_separator_rows_split_candidate_regions(tmp_path):
    path = tmp_path / "row-gaps.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in (1, 2, 4, 5):
        sheet.cell(row, 1, f"A{row}")
        sheet.cell(row, 2, f"B{row}")
    _save(workbook, path)

    result = profile_workbook(path).sheets[0]
    assert [run.model_dump() for run in result.blank_row_runs] == [{"start": 3, "end": 3}]
    assert [region.range for region in result.candidate_regions] == ["A1:B2", "A4:B5"]


def test_side_by_side_content_is_split_by_blank_column(tmp_path):
    path = tmp_path / "side-by-side.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in range(1, 6):
        for column in (1, 2, 3, 5, 6, 7):
            sheet.cell(row, column, f"{row}:{column}")
    _save(workbook, path)

    result = profile_workbook(path).sheets[0]
    assert [run.model_dump() for run in result.blank_column_runs] == [{"start": 4, "end": 4}]
    assert [region.range for region in result.candidate_regions] == ["A1:C5", "E1:G5"]


def test_hidden_rows_columns_and_freeze_panes_are_metadata_not_deletions(tmp_path):
    path = tmp_path / "hidden.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Visible"
    sheet["C4"] = "Hidden but present"
    sheet.row_dimensions[4].hidden = True
    sheet.column_dimensions["C"].hidden = True
    sheet.freeze_panes = "B2"
    _save(workbook, path)

    result = profile_workbook(path).sheets[0]
    assert result.hidden_rows == [4]
    assert result.hidden_columns == ["C"]
    assert result.freeze_panes == "B2"
    assert _cell(result, "C4").value == "Hidden but present"


def test_styled_empty_cell_affects_openpyxl_dimensions_not_content_bounds(tmp_path):
    path = tmp_path / "styled-empty.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Data"
    sheet["Z100"].font = Font(bold=True)
    _save(workbook, path)

    result = profile_workbook(path).sheets[0]
    assert result.openpyxl_max_row == 100
    assert result.openpyxl_max_column == 26
    assert result.content_range == "A1:A1"
    assert result.non_empty_cell_count == 1
    assert [cell.coordinate for cell in result.cells] == ["A1"]


def test_extreme_style_only_dimension_is_profiled_sparsely():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Real Data"
    sheet["XFD1048576"].font = Font(bold=True)

    result = profile_sheet(sheet)

    assert result.openpyxl_max_row == 1_048_576
    assert result.openpyxl_max_column == 16_384
    assert result.content_range == "A1:A1"
    assert result.non_empty_cell_count == 1
    assert [cell.coordinate for cell in result.cells] == ["A1"]
    workbook.close()


def test_profile_sheet_does_not_use_dense_iter_rows(monkeypatch):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = "Data"
    sheet["Z100"].font = Font(italic=True)

    def fail_dense_iteration(*args, **kwargs):
        raise AssertionError("profile_sheet must not call worksheet.iter_rows")

    monkeypatch.setattr(sheet, "iter_rows", fail_dense_iteration)
    result = profile_sheet(sheet)

    assert [cell.coordinate for cell in result.cells] == ["A1"]
    workbook.close()


def test_sparse_cell_order_is_row_major_not_creation_order():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["D10"] = "created first"
    sheet["A1"] = "created second"
    sheet["C4"] = "created third"

    result = profile_sheet(sheet)

    assert [cell.coordinate for cell in result.cells] == ["A1", "C4", "D10"]
    workbook.close()


def test_sparse_iterator_fails_closed_when_private_store_is_unavailable():
    class UnsupportedWorksheet:
        pass

    try:
        list(_iter_instantiated_cells(UnsupportedWorksheet()))  # type: ignore[arg-type]
    except WorkbookProfilerCompatibilityError as exc:
        assert "sparse cell access is unavailable" in str(exc)
    else:
        raise AssertionError("missing sparse storage must fail closed")


def test_empty_sheet_is_safe(tmp_path):
    path = tmp_path / "empty.xlsx"
    _save(openpyxl.Workbook(), path)

    result = profile_workbook(path).sheets[0]
    assert result.content_range is None
    assert result.content_min_row is None
    assert result.content_max_row is None
    assert result.content_min_column is None
    assert result.content_max_column is None
    assert result.non_empty_cell_count == 0
    assert result.cells == []
    assert result.row_profiles == []
    assert result.column_profiles == []
    assert result.candidate_regions == []


def test_formula_and_stable_style_signals_are_preserved_without_evaluation(tmp_path):
    path = tmp_path / "formula-style.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    cell = sheet["B2"]
    cell.value = "=SUM(B3:B4)"
    cell.font = Font(bold=True, italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    cell.fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    cell.number_format = "0.00"
    cell.border = Border(bottom=Side(style="thin"))
    sheet["B3"] = 2
    sheet["B4"] = 3
    _save(workbook, path)

    result = profile_workbook(path).sheets[0]
    formula = _cell(result, "B2")
    assert formula.is_formula is True
    assert formula.value_type == "formula"
    assert formula.value == "=SUM(B3:B4)"
    assert formula.style.bold is True
    assert formula.style.italic is True
    assert formula.style.horizontal_alignment == "center"
    assert formula.style.vertical_alignment == "top"
    assert formula.style.wrap_text is True
    assert formula.style.fill_type == "solid"
    assert formula.style.fill_color == "rgb:FFFF0000"
    assert formula.style.number_format == "0.00"
    assert formula.style.has_border is True
    assert result.row_profiles[0].formula_count == 1


def test_temporal_values_are_iso_serialized_with_explicit_types(tmp_path):
    path = tmp_path / "temporal.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet["A1"] = date(2026, 9, 5)
    sheet["B1"] = datetime(2026, 9, 5, 12, 30, 45)
    sheet["C1"] = time(8, 15, 0)
    _save(workbook, path)

    result = profile_workbook(path).sheets[0]
    assert _cell(result, "A1").value_type in {"date", "datetime"}
    assert _cell(result, "A1").value.startswith("2026-09-05")
    assert _cell(result, "B1").value_type == "datetime"
    assert _cell(result, "B1").value == "2026-09-05T12:30:45"
    assert _cell(result, "C1").value_type == "time"
    assert _cell(result, "C1").value == "08:15:00"


def test_all_worksheets_are_profiled_in_order_and_output_is_deterministic(tmp_path):
    path = tmp_path / "multi-sheet.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.title = "First"
    workbook.active["A1"] = "one"
    second = workbook.create_sheet("Second")
    second["D7"] = "two"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "three"
    _save(workbook, path)

    first = profile_workbook(path)
    second_profile = profile_workbook(path)

    assert first.workbook_sheet_names == ["First", "Second", "Hidden"]
    assert [sheet.sheet_name for sheet in first.sheets] == ["First", "Second", "Hidden"]
    assert first.sheets[2].sheet_state == "hidden"
    assert first.active_sheet == "First"
    assert first.source_file_sha256 == source_file_sha256(path)
    assert first.model_dump(mode="json") == second_profile.model_dump(mode="json")
