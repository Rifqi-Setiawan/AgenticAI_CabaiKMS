import openpyxl

from src.ingestion.structure_verifier import verify_structure
from src.ingestion.workbook_profiler import profile_workbook
from src.schema.structure import StructureProposal


def _profile(tmp_path):
    path = tmp_path / "structure.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Messy"
    ws.append(["Report title"])
    ws.append([])
    ws.append([None, "Morphology", None, "Fruit"])
    ws.append([None, "Variety", "Height", "Length"])
    ws.append([None, "Domba", 80, 5])
    ws.append([None, "Gendot", 90, 6])
    ws.merge_cells("B3:C3")
    wb.save(path)
    wb.close()
    return profile_workbook(path), path


def _row_proposal(**changes):
    row = {
        "table_range": "B3:D6",
        "header_rows": [3, 4],
        "data_start_row": 5,
        "data_end_row": 6,
        "attribute_columns": ["B", "C", "D"],
        "header_bindings": [
            {"column_letter": "B", "header_cells": ["B3", "B4"]},
            {"column_letter": "C", "header_cells": ["B3", "C4"]},
            {"column_letter": "D", "header_cells": ["D3", "D4"]},
        ],
    }
    row.update(changes)
    return StructureProposal(
        status="RESOLVED",
        orientation="row-oriented",
        row_oriented=row,
        confidence=0.9,
        reason_codes=["MULTILEVEL_HEADER"],
        evidence_summary="Header rows 3-4; data rows 5-6.",
    )


def test_valid_row_structure_and_merged_anchor(tmp_path):
    profile, _ = _profile(tmp_path)
    result = verify_structure(profile.sheets[0], _row_proposal())
    assert result.valid
    assert result.verified_orientation == "row-oriented"


def test_invalid_header_coordinate_fails_closed(tmp_path):
    profile, _ = _profile(tmp_path)
    proposal = _row_proposal(
        header_bindings=[
            {"column_letter": "B", "header_cells": ["B3", "B4"]},
            {"column_letter": "C", "header_cells": ["Z999"]},
            {"column_letter": "D", "header_cells": ["D3", "D4"]},
        ]
    )
    result = verify_structure(profile.sheets[0], proposal)
    assert not result.valid
    assert "HEADER_COORDINATE_OUTSIDE_TABLE" in result.issue_codes
    assert result.verified_orientation is None


def test_valid_transposed_structure(tmp_path):
    path = tmp_path / "transposed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T"
    ws.append(["Title"])
    ws.append([])
    ws.append([None, "Attribute", "Domba", "Gendot"])
    ws.append([None, "Habit", "terna", "perdu"])
    ws.append([None, "Height", 80, 90])
    wb.save(path)
    wb.close()
    sheet = profile_workbook(path).sheets[0]
    proposal = StructureProposal(
        status="RESOLVED",
        orientation="transposed",
        transposed={
            "table_range": "B3:D5",
            "header_row": 3,
            "label_column": "B",
            "data_columns": ["C", "D"],
            "attribute_start_row": 4,
            "attribute_end_row": 5,
        },
        confidence=0.8,
        evidence_summary="Transposed table.",
    )
    assert verify_structure(sheet, proposal).valid
