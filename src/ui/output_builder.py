"""Fase 8 — builds the actual "hasil akhir" output: a real instance of the
canonical template's shape (data/canonical/template_kanonik.xlsx) — same
Nomor/Karakter columns, same row labels, read dynamically from the template
(never hardcoded), but with varietas COLUMNS determined from the uploaded
source data itself (the anchor column's distinct values for row-oriented
input, or the transposed file's own column headers) rather than the
template's original 10 reference varieties.

A canonical row with no source attribute that mapped to it is simply left
blank — never guessed, never filled with a placeholder. Vision results are
written onto the SAME worksheet via src/agents/tabular_update.py exactly as
that module already does (Gambar row × matched_variety column, KNOWN status
only) — not reimplemented here.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from src.schema.canonical import DEFAULT_TEMPLATE_PATH, CanonicalSchema

SHEET_NAME = "Sheet1"
MULTI_VALUE_SEPARATOR = "; "


@dataclass
class CanonicalOutputBuilder:
    """Accumulates (canonical_row, varietas) -> normalized value pairs from
    the schema-matching pass, then materializes them into a workbook shaped
    exactly like the template."""

    schema: CanonicalSchema
    variety_names: list[str] = field(default_factory=list)
    _cells: dict[tuple[str, str], str] = field(default_factory=dict)

    def add_variety(self, name: str) -> None:
        if name not in self.variety_names:
            self.variety_names.append(name)

    def set_cell(self, row_id: str, variety_name: str, value: str | None) -> None:
        """If this cell already has a value — e.g. two different source
        attributes both mapped to the same canonical row for this variety —
        the new value is APPENDED with the project's multi-value separator
        rather than silently overwriting it, matching the same convention
        src/agents/tabular_update.py already uses for image references."""
        if value is None or str(value).strip() == "":
            return
        self.add_variety(variety_name)
        existing = self._cells.get((row_id, variety_name))
        if existing is None:
            self._cells[(row_id, variety_name)] = value
        else:
            existing_parts = [p.strip() for p in existing.split(MULTI_VALUE_SEPARATOR)]
            if value not in existing_parts:
                self._cells[(row_id, variety_name)] = existing + MULTI_VALUE_SEPARATOR + value

    def build_workbook(self, template_path: Path | str = DEFAULT_TEMPLATE_PATH) -> openpyxl.Workbook:
        wb = openpyxl.load_workbook(template_path)
        ws = wb[SHEET_NAME]

        n_rows = len(self.schema.rows)
        max_col = max(ws.max_column, 2 + len(self.variety_names))
        for col in range(3, max_col + 1):
            for row in range(1, n_rows + 2):
                # openpyxl's cell(..., value=None) is a no-op by design
                # (`if value is not None: cell.value = value`) — it does
                # NOT clear a cell, it just returns it unchanged. Must
                # assign .value directly to actually blank it out.
                ws.cell(row=row, column=col).value = None

        for i, name in enumerate(self.variety_names):
            ws.cell(row=1, column=3 + i, value=name)

        row_number_by_id = {row.id: idx + 2 for idx, row in enumerate(self.schema.rows)}
        for (row_id, variety), value in self._cells.items():
            if variety not in self.variety_names:
                continue
            r = row_number_by_id.get(row_id)
            if r is None:
                continue
            c = 3 + self.variety_names.index(variety)
            ws.cell(row=r, column=c, value=value)

        return wb


def worksheet_to_dataframe(ws: Worksheet, schema: CanonicalSchema, variety_names: list[str]) -> pd.DataFrame:
    """Read the actual current cell contents back out (post schema-matching
    AND post vision writes) — the single source of truth for what the
    preview table / Excel download show, rather than tracking two separate
    representations that could drift apart."""
    records = []
    for idx, row in enumerate(schema.rows):
        r = idx + 2
        record = {"Nomor": idx + 1, "Karakter": row.label}
        for i, variety in enumerate(variety_names):
            record[variety] = ws.cell(row=r, column=3 + i).value or ""
        records.append(record)
    return pd.DataFrame(records)


def values_by_variety(
    attr_row_values: list[str | None], position_to_variety: list[str | None]
) -> dict[str, list[str]]:
    """Groups one attribute's row_values by whichever varietas each
    position belongs to. `position_to_variety[i]` is the anchor column's
    row_values (row-oriented) or the transposed file's variety_names
    (transposed) — either way, index i in both lists refers to the same
    source row/column, which is exactly what ParsedAttribute.row_values'
    docstring promises stays aligned."""
    grouped: dict[str, list[str]] = {}
    for value, variety in zip(attr_row_values, position_to_variety):
        if value is None or variety is None:
            continue
        grouped.setdefault(variety, []).append(value)
    return grouped


def combine_multi_value(values: list[str]) -> str | None:
    """Multiple raw values for the same (canonical row, varietas) pair —
    e.g. several samples of the same varietas in a row-oriented file — are
    joined with the project's established multi-value separator (see
    src/agents/schema_matching/normalize.py) rather than picking just one
    arbitrarily and discarding the rest."""
    seen: list[str] = []
    for v in values:
        if v is not None and v not in seen:
            seen.append(v)
    if not seen:
        return None
    return MULTI_VALUE_SEPARATOR.join(seen)
