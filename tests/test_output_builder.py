from __future__ import annotations

from src.agents.tabular_update import apply_vision_result_to_worksheet
from src.schema.canonical import CanonicalRow, CanonicalSchema
from src.schema.contracts import ImageMetadata, VisionResult
from src.ui.output_builder import (
    SHEET_NAME,
    CanonicalOutputBuilder,
    combine_multi_value,
    values_by_variety,
    worksheet_to_dataframe,
)


def _tiny_schema() -> CanonicalSchema:
    rows = [
        CanonicalRow(id="r_1", label="habitus", domain="vegetatif", contoh_nilai=("perdu", "terna")),
        CanonicalRow(id="r_7", label="warna daun", domain="daun"),
        CanonicalRow(id="r_56", label="Lokasi", domain="lokasi"),
        CanonicalRow(id="r_57", label="Gambar Daun", domain="daun"),
    ]
    return CanonicalSchema(rows=rows, template_hash="test-hash", template_path=None)  # type: ignore[arg-type]


class TestValuesByVariety:
    def test_groups_row_oriented_style_by_position(self):
        attr_values = ["perdu", "terna", "terna", None]
        anchor_values = ["Gendot", "Kopay", "Kopay", "Katokkon"]
        grouped = values_by_variety(attr_values, anchor_values)
        assert grouped == {"Gendot": ["perdu"], "Kopay": ["terna", "terna"]}

    def test_transposed_style_one_to_one_alignment(self):
        attr_values = ["perdu", "terna", "terna"]
        variety_names = ["Gendot", "Kopay", "Katokkon"]
        grouped = values_by_variety(attr_values, variety_names)
        assert grouped == {"Gendot": ["perdu"], "Kopay": ["terna"], "Katokkon": ["terna"]}

    def test_none_position_or_value_is_skipped(self):
        assert values_by_variety([None, "x"], ["A", None]) == {}
        assert values_by_variety(["x", None], ["A", "B"]) == {"A": ["x"]}


class TestCombineMultiValue:
    def test_single_value_passthrough(self):
        assert combine_multi_value(["perdu"]) == "perdu"

    def test_multiple_distinct_values_joined(self):
        assert combine_multi_value(["perdu", "terna"]) == "perdu; terna"

    def test_duplicate_values_deduplicated(self):
        assert combine_multi_value(["terna", "terna", "terna"]) == "terna"

    def test_empty_list_returns_none(self):
        assert combine_multi_value([]) is None

    def test_all_none_returns_none(self):
        assert combine_multi_value([None, None]) is None  # type: ignore[list-item]


class TestCanonicalOutputBuilderSetCell:
    def test_first_write_sets_value_and_registers_variety(self):
        builder = CanonicalOutputBuilder(schema=_tiny_schema())
        builder.set_cell("r_1", "Gendot", "perdu")
        assert builder.variety_names == ["Gendot"]

    def test_none_or_blank_value_is_ignored(self):
        builder = CanonicalOutputBuilder(schema=_tiny_schema())
        builder.set_cell("r_1", "Gendot", None)
        builder.set_cell("r_1", "Gendot", "   ")
        assert builder.variety_names == []

    def test_second_distinct_write_to_same_cell_appends(self):
        """Two different source attributes mapping to the same canonical
        row for the same variety must merge, not silently overwrite —
        matching src/agents/tabular_update.py's own convention."""
        builder = CanonicalOutputBuilder(schema=_tiny_schema())
        builder.set_cell("r_56", "Gendot", "Lokasi ke-1")
        builder.set_cell("r_56", "Gendot", "Lahan Percobaan 1, Indonesia")
        assert builder._cells[("r_56", "Gendot")] == "Lokasi ke-1; Lahan Percobaan 1, Indonesia"

    def test_rewriting_the_identical_value_does_not_duplicate(self):
        builder = CanonicalOutputBuilder(schema=_tiny_schema())
        builder.set_cell("r_1", "Gendot", "perdu")
        builder.set_cell("r_1", "Gendot", "perdu")
        assert builder._cells[("r_1", "Gendot")] == "perdu"


class TestBuildWorkbookAndDataframe:
    def test_workbook_has_dynamic_row_labels_and_source_derived_columns(self, tmp_path):
        schema = _tiny_schema()
        # a minimal template file matching the tiny schema's shape, so
        # build_workbook has something real to load and clear
        import openpyxl

        template_path = tmp_path / "template.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["Nomor", "Karakter", "OldRefVarietyA", "OldRefVarietyB"])
        ws.append([1, "habitus", "perdu", "terna"])
        ws.append([2, "warna daun", "green", "green"])
        ws.append([3, "Lokasi", None, None])
        ws.append([4, "Gambar Daun", None, None])
        wb.save(template_path)

        builder = CanonicalOutputBuilder(schema=schema)
        builder.set_cell("r_1", "Gendot", "perdu")
        builder.set_cell("r_7", "Gendot", "green group 137 A")
        builder.set_cell("r_1", "Kopay", "terna")

        out_wb = builder.build_workbook(template_path=template_path)
        ws_out = out_wb[SHEET_NAME]

        assert ws_out.cell(row=1, column=3).value == "Gendot"
        assert ws_out.cell(row=1, column=4).value == "Kopay"
        assert ws_out.cell(row=2, column=3).value == "perdu"  # r_1 x Gendot
        assert ws_out.cell(row=3, column=3).value == "green group 137 A"  # r_7 x Gendot
        assert ws_out.cell(row=2, column=4).value == "terna"  # r_1 x Kopay
        assert ws_out.cell(row=4, column=3).value is None  # r_56 x Gendot — never mapped, stays blank

    def test_old_reference_columns_are_fully_cleared_not_leaked(self, tmp_path):
        """The template's own original varietas columns (headers AND
        values) must not leak into the output when the source data
        determines a different, smaller set of varieties."""
        import openpyxl

        template_path = tmp_path / "template.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["Nomor", "Karakter"] + [f"RefVariety{i}" for i in range(10)])
        ws.append([1, "habitus"] + ["perdu"] * 10)
        wb.save(template_path)

        schema = CanonicalSchema(
            rows=[CanonicalRow(id="r_1", label="habitus", domain="vegetatif")],
            template_hash="h", template_path=None,  # type: ignore[arg-type]
        )
        builder = CanonicalOutputBuilder(schema=schema)
        builder.set_cell("r_1", "Gendot", "perdu")

        out_wb = builder.build_workbook(template_path=template_path)
        ws_out = out_wb[SHEET_NAME]

        assert ws_out.cell(row=1, column=3).value == "Gendot"
        assert ws_out.cell(row=1, column=4).value is None  # old RefVariety1 header gone
        assert ws_out.cell(row=2, column=4).value is None  # old RefVariety1 value gone

    def test_worksheet_to_dataframe_reflects_current_cell_contents(self, tmp_path):
        import openpyxl

        template_path = tmp_path / "template.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["Nomor", "Karakter"])
        ws.append([1, "habitus"])
        ws.append([2, "warna daun"])
        wb.save(template_path)

        schema = CanonicalSchema(
            rows=[
                CanonicalRow(id="r_1", label="habitus", domain="vegetatif"),
                CanonicalRow(id="r_7", label="warna daun", domain="daun"),
            ],
            template_hash="h", template_path=None,  # type: ignore[arg-type]
        )
        builder = CanonicalOutputBuilder(schema=schema)
        builder.set_cell("r_1", "Gendot", "perdu")
        out_wb = builder.build_workbook(template_path=template_path)
        ws_out = out_wb[SHEET_NAME]

        df = worksheet_to_dataframe(ws_out, schema, builder.variety_names)
        assert df.loc[df["Karakter"] == "habitus", "Gendot"].iloc[0] == "perdu"
        assert df.loc[df["Karakter"] == "warna daun", "Gendot"].iloc[0] == ""

    def test_vision_result_written_via_tabular_update_is_reflected_in_dataframe(self, tmp_path):
        """Integration check: build_workbook's output is a valid target for
        Fase 6's apply_vision_result_to_worksheet, and worksheet_to_dataframe
        picks up that write afterward — this is exactly the sequencing
        src/ui/pipeline_runner.py relies on."""
        import openpyxl

        template_path = tmp_path / "template.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["Nomor", "Karakter"])
        ws.append([1, "Gambar Daun"])
        wb.save(template_path)

        schema = CanonicalSchema(
            rows=[CanonicalRow(id="r_1", label="Gambar Daun", domain="daun")],
            template_hash="h", template_path=None,  # type: ignore[arg-type]
        )
        builder = CanonicalOutputBuilder(schema=schema, variety_names=["Gendot"])
        out_wb = builder.build_workbook(template_path=template_path)
        ws_out = out_wb[SHEET_NAME]

        image = ImageMetadata(
            file_id="abc123", filename="daun.jpg", mime_type="image/jpeg", size=100,
            created_time="2026-01-01T00:00:00Z",
        )
        vision_result = VisionResult(
            classification_status="KNOWN", matched_variety="Gendot", identified_part="DAUN",
            confidence=0.9, visual_evidence="ok",
        )
        update_result = apply_vision_result_to_worksheet(ws_out, image, vision_result)
        assert update_result.applied is True

        df = worksheet_to_dataframe(ws_out, schema, builder.variety_names)
        assert "abc123" in df.loc[df["Karakter"] == "Gambar Daun", "Gendot"].iloc[0]
