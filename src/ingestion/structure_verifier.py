"""Deterministic verification of probabilistic worksheet structure proposals."""

from __future__ import annotations

from openpyxl.utils import column_index_from_string

from src.ingestion.structure_geometry import (
    StructureGeometryError,
    cell_lookup,
    content_bounds,
    parse_cell_coordinate,
    parse_excel_range,
    resolve_profile_cell,
)
from src.ingestion.workbook_profiler import SheetProfile
from src.schema.structure import StructureProposal, StructureStatus, StructureVerificationResult


def verify_structure(sheet: SheetProfile, proposal: StructureProposal) -> StructureVerificationResult:
    issues: list[tuple[str, str]] = []

    def issue(code: str, detail: str) -> None:
        issues.append((code, detail))

    if proposal.status is not StructureStatus.RESOLVED:
        issue("STATUS_NOT_RESOLVED", f"proposal status is {proposal.status.value}")
        return _result(issues, None)

    structure = proposal.row_oriented or proposal.transposed
    try:
        table = parse_excel_range(structure.table_range, expected_sheet=sheet.sheet_name)  # type: ignore[union-attr]
    except StructureGeometryError as exc:
        issue("INVALID_TABLE_RANGE", str(exc))
        return _result(issues, proposal.orientation)

    actual = content_bounds(sheet)
    if actual is None:
        issue("EMPTY_SHEET", "resolved structure cannot target an empty sheet")
        return _result(issues, proposal.orientation)
    if not table.intersects(actual):
        issue("TABLE_OUTSIDE_CONTENT", "table range does not intersect actual content")
    if not (
        actual.min_row <= table.min_row <= table.max_row <= actual.max_row
        and actual.min_column <= table.min_column <= table.max_column <= actual.max_column
    ):
        issue("TABLE_NOT_WITHIN_CONTENT_BOUNDS", "table range exceeds actual content bounds")
    if not any(table.contains(cell.row, cell.column) for cell in sheet.cells):
        issue("TABLE_HAS_NO_CONTENT", "table range contains no profiled content")

    if proposal.orientation == "row-oriented":
        _verify_row_oriented(sheet, proposal, table, issue)
    else:
        _verify_transposed(sheet, proposal, table, issue)
    return _result(issues, proposal.orientation)


def _result(issues: list[tuple[str, str]], orientation) -> StructureVerificationResult:
    return StructureVerificationResult(
        valid=not issues,
        issue_codes=[code for code, _ in issues],
        issue_details=[detail for _, detail in issues],
        verified_orientation=orientation if not issues else None,
    )


def _column_index(value: str, issue, code: str) -> int | None:
    try:
        return column_index_from_string(value.replace("$", "").upper())
    except ValueError:
        issue(code, f"invalid column letter: {value!r}")
        return None


def _verify_row_oriented(sheet, proposal, table, issue) -> None:
    structure = proposal.row_oriented
    assert structure is not None
    headers = structure.header_rows
    if not headers:
        issue("HEADER_ROWS_EMPTY", "header_rows must be non-empty")
    if headers != sorted(set(headers)):
        issue("HEADER_ROWS_INVALID_ORDER", "header_rows must be sorted and unique")
    for row in headers:
        if not table.min_row <= row <= table.max_row:
            issue("HEADER_ROW_OUTSIDE_TABLE", f"header row {row} is outside table")
    if headers and structure.data_start_row <= max(headers):
        issue("DATA_BEFORE_HEADER_END", "data_start_row must be after all header rows")
    if structure.data_end_row < structure.data_start_row:
        issue("INVALID_DATA_ROWS", "data_end_row precedes data_start_row")
    if not (
        table.min_row <= structure.data_start_row <= table.max_row
        and table.min_row <= structure.data_end_row <= table.max_row
    ):
        issue("DATA_ROWS_OUTSIDE_TABLE", "data rows are outside table range")

    normalized_columns = [column.replace("$", "").upper() for column in structure.attribute_columns]
    if not normalized_columns or len(normalized_columns) != len(set(normalized_columns)):
        issue("ATTRIBUTE_COLUMNS_INVALID", "attribute_columns must be non-empty and unique")
    column_indices: dict[str, int] = {}
    for column in normalized_columns:
        index = _column_index(column, issue, "INVALID_ATTRIBUTE_COLUMN")
        if index is not None:
            column_indices[column] = index
            if not table.min_column <= index <= table.max_column:
                issue("ATTRIBUTE_COLUMN_OUTSIDE_TABLE", f"attribute column {column} is outside table")

    binding_columns = [binding.column_letter.replace("$", "").upper() for binding in structure.header_bindings]
    if len(binding_columns) != len(set(binding_columns)) or set(binding_columns) != set(normalized_columns):
        issue("HEADER_BINDING_MISMATCH", "exactly one header binding is required per attribute column")
    lookup = cell_lookup(sheet)
    for binding in structure.header_bindings:
        if not binding.header_cells:
            issue("HEADER_BINDING_EMPTY", f"binding {binding.column_letter!r} has no header cells")
            continue
        resolved_values = []
        for coordinate in binding.header_cells:
            try:
                row, column = parse_cell_coordinate(coordinate)
            except StructureGeometryError as exc:
                issue("INVALID_HEADER_COORDINATE", str(exc))
                continue
            if not table.contains(row, column):
                issue("HEADER_COORDINATE_OUTSIDE_TABLE", f"{coordinate} is outside table")
            if row not in headers:
                issue("HEADER_COORDINATE_WRONG_ROW", f"{coordinate} is not on a declared header row")
            resolved = resolve_profile_cell(sheet, coordinate, cells_by_coordinate=lookup)
            if resolved is None:
                issue("HEADER_CELL_HAS_NO_SOURCE", f"{coordinate} has no content or merged anchor")
            else:
                resolved_values.append(resolved[1].value)
        if not resolved_values or str(resolved_values[-1]).strip() == "":
            issue("LEAF_HEADER_UNUSABLE", f"binding {binding.column_letter!r} has no usable leaf label")


def _verify_transposed(sheet, proposal, table, issue) -> None:
    structure = proposal.transposed
    assert structure is not None
    if not table.min_row <= structure.header_row <= table.max_row:
        issue("HEADER_ROW_OUTSIDE_TABLE", "header_row is outside table")
    label_index = _column_index(structure.label_column, issue, "INVALID_LABEL_COLUMN")
    if label_index is not None and not table.min_column <= label_index <= table.max_column:
        issue("LABEL_COLUMN_OUTSIDE_TABLE", "label_column is outside table")
    normalized_data = [column.replace("$", "").upper() for column in structure.data_columns]
    if not normalized_data or len(normalized_data) != len(set(normalized_data)):
        issue("DATA_COLUMNS_INVALID", "data_columns must be non-empty and unique")
    data_indices: list[tuple[str, int]] = []
    for column in normalized_data:
        index = _column_index(column, issue, "INVALID_DATA_COLUMN")
        if index is not None:
            data_indices.append((column, index))
            if not table.min_column <= index <= table.max_column:
                issue("DATA_COLUMN_OUTSIDE_TABLE", f"data column {column} is outside table")
            if label_index == index:
                issue("DATA_COLUMN_EQUALS_LABEL", f"data column {column} equals label column")
    if structure.attribute_start_row <= structure.header_row:
        issue("ATTRIBUTES_BEFORE_HEADER", "attribute_start_row must be after header_row")
    if structure.attribute_end_row < structure.attribute_start_row:
        issue("INVALID_ATTRIBUTE_ROWS", "attribute_end_row precedes attribute_start_row")
    if not (
        table.min_row <= structure.attribute_start_row <= table.max_row
        and table.min_row <= structure.attribute_end_row <= table.max_row
    ):
        issue("ATTRIBUTE_ROWS_OUTSIDE_TABLE", "attribute rows are outside table")

    lookup = cell_lookup(sheet)
    for column, _ in data_indices:
        coordinate = f"{column}{structure.header_row}"
        if resolve_profile_cell(sheet, coordinate, cells_by_coordinate=lookup) is None:
            issue("ENTITY_HEADER_MISSING", f"{coordinate} has no entity header content")
    if label_index is not None:
        has_label = any(
            resolve_profile_cell(
                sheet,
                f"{structure.label_column.upper()}{row}",
                cells_by_coordinate=lookup,
            )
            is not None
            for row in range(structure.attribute_start_row, structure.attribute_end_row + 1)
        )
        if not has_label:
            issue("ATTRIBUTE_LABELS_MISSING", "label column has no attribute labels in declared rows")
