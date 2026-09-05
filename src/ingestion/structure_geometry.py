"""Shared deterministic Excel geometry and merged-anchor resolution helpers."""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from src.ingestion.workbook_profiler import CellProfile, SheetProfile

EXCEL_MAX_ROW = 1_048_576
EXCEL_MAX_COLUMN = 16_384


class StructureGeometryError(ValueError):
    pass


@dataclass(frozen=True)
class RangeBounds:
    min_column: int
    min_row: int
    max_column: int
    max_row: int

    @property
    def cell_count(self) -> int:
        return (self.max_row - self.min_row + 1) * (self.max_column - self.min_column + 1)

    def contains(self, row: int, column: int) -> bool:
        return self.min_row <= row <= self.max_row and self.min_column <= column <= self.max_column

    def intersects(self, other: RangeBounds) -> bool:
        return not (
            self.max_row < other.min_row
            or self.min_row > other.max_row
            or self.max_column < other.min_column
            or self.min_column > other.max_column
        )

    @property
    def coordinate(self) -> str:
        return (
            f"{get_column_letter(self.min_column)}{self.min_row}:"
            f"{get_column_letter(self.max_column)}{self.max_row}"
        )


def parse_excel_range(reference: str, *, expected_sheet: str | None = None) -> RangeBounds:
    text = reference.strip()
    if "!" in text:
        sheet_part, text = text.rsplit("!", 1)
        sheet_part = sheet_part.strip("'").replace("''", "'")
        if expected_sheet is not None and sheet_part != expected_sheet:
            raise StructureGeometryError(f"range refers to sheet {sheet_part!r}, expected {expected_sheet!r}")
    text = text.replace("$", "")
    if ":" not in text:
        text = f"{text}:{text}"
    try:
        min_column, min_row, max_column, max_row = range_boundaries(text)
    except (TypeError, ValueError) as exc:
        raise StructureGeometryError(f"invalid Excel range: {reference!r}") from exc
    if not all(isinstance(value, int) for value in (min_column, min_row, max_column, max_row)):
        raise StructureGeometryError(f"range must use concrete cell coordinates: {reference!r}")
    if (
        min_row < 1
        or min_column < 1
        or max_row > EXCEL_MAX_ROW
        or max_column > EXCEL_MAX_COLUMN
        or min_row > max_row
        or min_column > max_column
    ):
        raise StructureGeometryError(f"range exceeds Excel limits: {reference!r}")
    return RangeBounds(min_column, min_row, max_column, max_row)


def parse_cell_coordinate(coordinate: str) -> tuple[int, int]:
    try:
        row, column = coordinate_to_tuple(coordinate.replace("$", ""))
    except (TypeError, ValueError) as exc:
        raise StructureGeometryError(f"invalid cell coordinate: {coordinate!r}") from exc
    if row > EXCEL_MAX_ROW or column > EXCEL_MAX_COLUMN:
        raise StructureGeometryError(f"cell coordinate exceeds Excel limits: {coordinate!r}")
    return row, column


def content_bounds(sheet: SheetProfile) -> RangeBounds | None:
    if sheet.content_range is None:
        return None
    return parse_excel_range(sheet.content_range, expected_sheet=sheet.sheet_name)


def cell_lookup(sheet: SheetProfile) -> dict[str, CellProfile]:
    return {cell.coordinate: cell for cell in sheet.cells}


def resolve_profile_cell(
    sheet: SheetProfile,
    coordinate: str,
    *,
    cells_by_coordinate: dict[str, CellProfile] | None = None,
) -> tuple[str, CellProfile] | None:
    """Resolve direct content or a merged placeholder to its real top-left cell."""
    normalized = coordinate.replace("$", "").upper()
    lookup = cells_by_coordinate if cells_by_coordinate is not None else cell_lookup(sheet)
    direct = lookup.get(normalized)
    if direct is not None:
        return normalized, direct
    row, column = parse_cell_coordinate(normalized)
    for merged in sheet.merged_ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_column <= column <= merged.max_column:
            anchor = lookup.get(merged.top_left_coordinate)
            return (merged.top_left_coordinate, anchor) if anchor is not None else None
    return None
