import openpyxl

from src.agents.schema_matching.anchor import AnchorResult
from src.ingestion import shadow_pipeline
from src.ingestion.shadow_pipeline import run_structure_shadow
from src.schema.structure import StructureProposal


def _anchor(candidates, **kwargs):
    selected = next((item.column_name for item in candidates if item.column_name == "Variety"), None)
    return AnchorResult("found" if selected else "escalate", selected, 1.0, "test")


def _llm(proposal):
    return lambda **kwargs: proposal


def _flat(tmp_path):
    path = tmp_path / "flat.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Variety", "Height"])
    ws.append(["Domba", 80])
    ws.append(["Gendot", None])
    wb.save(path)
    wb.close()
    return path


def _flat_proposal():
    return StructureProposal(
        status="RESOLVED", orientation="row-oriented",
        row_oriented={
            "table_range": "A1:B3", "header_rows": [1],
            "data_start_row": 2, "data_end_row": 3,
            "attribute_columns": ["A", "B"],
            "header_bindings": [
                {"column_letter": "A", "header_cells": ["A1"]},
                {"column_letter": "B", "header_cells": ["B1"]},
            ],
        }, confidence=0.8, evidence_summary="Flat table.",
    )


def test_simple_row_shadow_matches_and_profiles_once(tmp_path, monkeypatch):
    path = _flat(tmp_path)
    real_profile = shadow_pipeline.profile_workbook
    calls = []

    def profile_once(file_path):
        calls.append(file_path)
        return real_profile(file_path)

    monkeypatch.setattr(shadow_pipeline, "profile_workbook", profile_once)
    report = run_structure_shadow(
        path, "Data", source_format="row-oriented",
        llm_call=_llm(_flat_proposal()), anchor_detector=_anchor,
    )
    assert report.status.value == "MATCH"
    assert len(calls) == 1


def test_two_row_header_shadow_matches(tmp_path):
    path = tmp_path / "two-row.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Identity", "Fruit", None])
    ws.append(["Variety", "Length", "Width"])
    ws.append(["Domba", "3", "2"])
    ws.append(["Gendot", "5", None])
    ws.merge_cells("B1:C1")
    wb.save(path)
    wb.close()
    proposal = StructureProposal(
        status="RESOLVED", orientation="row-oriented",
        row_oriented={
            "table_range": "A1:C4", "header_rows": [1, 2],
            "data_start_row": 3, "data_end_row": 4,
            "attribute_columns": ["A", "B", "C"],
            "header_bindings": [
                {"column_letter": "A", "header_cells": ["A1", "A2"]},
                {"column_letter": "B", "header_cells": ["B1", "B2"]},
                {"column_letter": "C", "header_cells": ["B1", "C2"]},
            ],
        }, confidence=0.8, evidence_summary="Two header rows.",
    )
    report = run_structure_shadow(
        path, "Data", source_format="row-oriented", header_rows=2,
        llm_call=_llm(proposal), anchor_detector=_anchor,
    )
    assert report.status.value == "MATCH"
    assert report.attribute_identity_parity == 1.0


def test_transposed_shadow_matches_entity_positions(tmp_path):
    path = tmp_path / "transposed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "T"
    ws.append(["Karakter", "Domba", "Gendot"])
    ws.append(["Habit", "terna", "perdu"])
    ws.append(["Height", 80, None])
    wb.save(path)
    wb.close()
    proposal = StructureProposal(
        status="RESOLVED", orientation="transposed",
        transposed={
            "table_range": "A1:C3", "header_row": 1, "label_column": "A",
            "data_columns": ["B", "C"], "attribute_start_row": 2,
            "attribute_end_row": 3,
        }, confidence=0.8, evidence_summary="Transposed table.",
    )
    report = run_structure_shadow(
        path, "T", source_format="transposed", llm_call=_llm(proposal)
    )
    assert report.status.value == "MATCH"
    assert report.entity_names_match
    assert report.value_position_parity == 1.0


def test_new_path_abstention_preserves_observable_legacy_output(tmp_path):
    report = run_structure_shadow(
        _flat(tmp_path), "Data", source_format="row-oriented",
        llm_call=_llm(StructureProposal(
            status="AMBIGUOUS", confidence=0.2,
            reason_codes=["INSUFFICIENT_EVIDENCE"], evidence_summary="Unclear.",
        )), anchor_detector=_anchor,
    )
    assert report.status.value == "NEW_PATH_ABSTAINED"
    assert report.legacy_attribute_count == 2
    assert report.structure_status == "AMBIGUOUS"


def test_invalid_new_structure_records_verifier_failure(tmp_path):
    bad = _flat_proposal().model_copy(deep=True)
    bad.row_oriented.header_bindings[1].header_cells = ["Z999"]
    report = run_structure_shadow(
        _flat(tmp_path), "Data", source_format="row-oriented",
        llm_call=_llm(bad), anchor_detector=_anchor,
    )
    assert report.status.value == "NEW_PATH_FAILED"
    assert "HEADER_COORDINATE_OUTSIDE_TABLE" in report.verification_issue_codes
    assert not report.new_path_resolved


def _messy(tmp_path):
    path = tmp_path / "messy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Messy"
    ws.append(["Research title"])
    ws.append(["Location", "Bogor"])
    ws.append([])
    ws.append([None, "Identity", "Morphology"])
    ws.append([None, "Variety", "Height"])
    ws.append([None, "name", "cm"])
    ws.append([None, "Domba", 80])
    ws.append([None, "Gendot", 90])
    wb.save(path)
    wb.close()
    return path


def _messy_proposal():
    return StructureProposal(
        status="RESOLVED", orientation="row-oriented",
        row_oriented={
            "table_range": "B4:C8", "header_rows": [4, 5, 6],
            "data_start_row": 7, "data_end_row": 8,
            "attribute_columns": ["B", "C"],
            "header_bindings": [
                {"column_letter": "B", "header_cells": ["B4", "B5"]},
                {"column_letter": "C", "header_cells": ["C4", "C5"]},
            ],
        }, confidence=0.9, evidence_summary="Table begins at row four.",
    )


def test_legacy_failure_new_success_is_coverage_signal_not_match(tmp_path):
    report = run_structure_shadow(
        _messy(tmp_path), "Messy", source_format="row-oriented",
        llm_call=_llm(_messy_proposal()), anchor_detector=_anchor,
    )
    assert report.status.value == "LEGACY_FAILED"
    assert report.new_path_resolved
    assert report.legacy_error_type == "ValueError"


def test_both_fail_returns_structured_report(tmp_path):
    report = run_structure_shadow(
        _messy(tmp_path), "Messy", source_format="row-oriented",
        llm_call=_llm(StructureProposal(
            status="UNSUPPORTED", confidence=0.1, evidence_summary="Not supported.",
        )), anchor_detector=_anchor,
    )
    assert report.status.value == "BOTH_FAILED"
    assert report.legacy_error_type == "ValueError"
    assert report.structure_status == "UNSUPPORTED"
