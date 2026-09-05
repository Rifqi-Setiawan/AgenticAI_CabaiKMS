"""Deterministic physical-structure profiling for Excel workbooks.

This module records observable workbook facts only. It deliberately does not
infer tables, headers, orientation, identifiers, or canonical mappings.
"""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Literal

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Color
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from src.schema.provenance import source_file_sha256

PROFILE_VERSION = "workbook-structure-v1"

CellValueType = Literal[
    "string", "integer", "float", "boolean", "date", "datetime", "time", "formula", "other"
]


class CellStyleProfile(BaseModel):
    style_id: int
    bold: bool
    italic: bool
    horizontal_alignment: str | None
    vertical_alignment: str | None
    wrap_text: bool | None
    fill_type: str | None
    fill_color: str | None
    number_format: str
    has_border: bool


class CellProfile(BaseModel):
    coordinate: str
    row: int
    column: int
    column_letter: str
    value: Any
    value_type: CellValueType
    is_formula: bool
    style: CellStyleProfile
    merged_range: str | None = None


class MergedRangeProfile(BaseModel):
    range: str
    min_row: int
    max_row: int
    min_column: int
    max_column: int
    top_left_coordinate: str
    top_left_value: Any = None


class IndexRunProfile(BaseModel):
    start: int
    end: int


class RowProfile(BaseModel):
    row_index: int
    non_empty_count: int
    text_count: int
    numeric_count: int
    boolean_count: int
    formula_count: int
    first_non_empty_column: int | None
    last_non_empty_column: int | None
    is_blank: bool


class ColumnProfile(BaseModel):
    column_index: int
    column_letter: str
    non_empty_count: int
    text_count: int
    numeric_count: int
    boolean_count: int
    formula_count: int
    first_non_empty_row: int | None
    last_non_empty_row: int | None
    is_blank: bool


class CandidateRegionProfile(BaseModel):
    range: str
    min_row: int
    max_row: int
    min_column: int
    max_column: int
    non_empty_cells: int
    total_cells: int
    density: float


class SheetProfile(BaseModel):
    sheet_name: str
    sheet_state: str
    openpyxl_max_row: int
    openpyxl_max_column: int
    content_min_row: int | None
    content_max_row: int | None
    content_min_column: int | None
    content_max_column: int | None
    content_range: str | None
    non_empty_cell_count: int
    merged_ranges: list[MergedRangeProfile]
    hidden_rows: list[int]
    hidden_columns: list[str]
    freeze_panes: str | None
    row_profiles: list[RowProfile]
    column_profiles: list[ColumnProfile]
    cells: list[CellProfile]
    blank_row_runs: list[IndexRunProfile]
    blank_column_runs: list[IndexRunProfile]
    candidate_regions: list[CandidateRegionProfile]


class WorkbookProfile(BaseModel):
    profile_version: Literal["workbook-structure-v1"] = PROFILE_VERSION
    source_file_name: str
    source_file_sha256: str
    workbook_sheet_names: list[str]
    sheets: list[SheetProfile]
    defined_names_count: int
    active_sheet: str | None


def _has_content(value: Any) -> bool:
    """The single content-presence rule used by all profiler calculations."""
    return value is not None and str(value).strip() != ""


def _value_type(cell: Cell) -> CellValueType:
    value = cell.value
    if cell.data_type == "f":
        return "formula"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, time):
        return "time"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "float"
    if isinstance(value, str):
        return "string"
    return "other"


def _serializable_value(value: Any, value_type: CellValueType | None = None) -> Any:
    """Retain scalar types; encode temporal values explicitly as ISO strings."""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if value_type == "other" and value is not None:
        return str(value)
    return value


def _serialize_color(color: Color | None) -> str | None:
    if color is None:
        return None
    try:
        color_type = color.type
        if color_type == "rgb" and color.rgb is not None:
            return f"rgb:{color.rgb}"
        if color_type == "theme" and color.theme is not None:
            return f"theme:{color.theme}"
        if color_type == "indexed" and color.indexed is not None:
            return f"indexed:{color.indexed}"
        if color_type == "auto" and color.auto is not None:
            return f"auto:{str(bool(color.auto)).lower()}"
    except (AttributeError, TypeError, ValueError):
        return None
    return None


def _has_border(cell: Cell) -> bool:
    sides = (
        cell.border.left,
        cell.border.right,
        cell.border.top,
        cell.border.bottom,
        cell.border.diagonal,
        cell.border.vertical,
        cell.border.horizontal,
    )
    return any(side is not None and side.style is not None for side in sides)


def _style_profile(cell: Cell) -> CellStyleProfile:
    fill_type = cell.fill.fill_type
    return CellStyleProfile(
        style_id=cell.style_id,
        bold=bool(cell.font.bold),
        italic=bool(cell.font.italic),
        horizontal_alignment=cell.alignment.horizontal,
        vertical_alignment=cell.alignment.vertical,
        wrap_text=cell.alignment.wrap_text,
        fill_type=fill_type,
        fill_color=_serialize_color(cell.fill.fgColor) if fill_type else None,
        number_format=cell.number_format,
        has_border=_has_border(cell),
    )


def _contiguous_runs(indices: list[int]) -> list[tuple[int, int]]:
    if not indices:
        return []
    ordered = sorted(set(indices))
    runs: list[tuple[int, int]] = []
    start = previous = ordered[0]
    for current in ordered[1:]:
        if current != previous + 1:
            runs.append((start, previous))
            start = current
        previous = current
    runs.append((start, previous))
    return runs


def _blank_runs(start: int, end: int, active: set[int]) -> list[IndexRunProfile]:
    blanks = [index for index in range(start, end + 1) if index not in active]
    return [IndexRunProfile(start=a, end=b) for a, b in _contiguous_runs(blanks)]


def _counts(cells: list[CellProfile]) -> tuple[int, int, int, int, int]:
    text = sum(cell.value_type == "string" for cell in cells)
    numeric = sum(cell.value_type in {"integer", "float"} for cell in cells)
    boolean = sum(cell.value_type == "boolean" for cell in cells)
    formula = sum(cell.value_type == "formula" for cell in cells)
    return len(cells), text, numeric, boolean, formula


def _candidate_regions(cells: list[CellProfile]) -> list[CandidateRegionProfile]:
    active_rows = sorted({cell.row for cell in cells})
    regions: list[CandidateRegionProfile] = []
    for min_row, max_row in _contiguous_runs(active_rows):
        band_cells = [cell for cell in cells if min_row <= cell.row <= max_row]
        active_columns = sorted({cell.column for cell in band_cells})
        for min_column, max_column in _contiguous_runs(active_columns):
            count = sum(min_column <= cell.column <= max_column for cell in band_cells)
            if count == 0:
                continue
            total = (max_row - min_row + 1) * (max_column - min_column + 1)
            regions.append(
                CandidateRegionProfile(
                    range=(
                        f"{get_column_letter(min_column)}{min_row}:"
                        f"{get_column_letter(max_column)}{max_row}"
                    ),
                    min_row=min_row,
                    max_row=max_row,
                    min_column=min_column,
                    max_column=max_column,
                    non_empty_cells=count,
                    total_cells=total,
                    density=(count / total) if total else 0.0,
                )
            )
    return regions


def _hidden_columns(worksheet: Worksheet) -> list[str]:
    hidden: set[int] = set()
    for key, dimension in worksheet.column_dimensions.items():
        if not dimension.hidden:
            continue
        first = dimension.min or column_index_from_string(key)
        last = dimension.max or first
        hidden.update(range(first, last + 1))
    return [get_column_letter(index) for index in sorted(hidden)]


def profile_sheet(worksheet: Worksheet) -> SheetProfile:
    """Profile one already-open worksheet without semantic interpretation."""
    merged_ranges: list[MergedRangeProfile] = []
    merged_anchor: dict[str, str] = {}
    ordered_merges = sorted(
        worksheet.merged_cells.ranges,
        key=lambda item: (item.min_row, item.min_col, item.max_row, item.max_col),
    )
    for merged in ordered_merges:
        range_text = str(merged)
        top_left = worksheet.cell(merged.min_row, merged.min_col)
        value_type = _value_type(top_left) if _has_content(top_left.value) else None
        merged_ranges.append(
            MergedRangeProfile(
                range=range_text,
                min_row=merged.min_row,
                max_row=merged.max_row,
                min_column=merged.min_col,
                max_column=merged.max_col,
                top_left_coordinate=top_left.coordinate,
                top_left_value=_serializable_value(top_left.value, value_type),
            )
        )
        merged_anchor[top_left.coordinate] = range_text

    cells: list[CellProfile] = []
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            if not _has_content(cell.value):
                continue
            value_type = _value_type(cell)
            cells.append(
                CellProfile(
                    coordinate=cell.coordinate,
                    row=cell.row,
                    column=cell.column,
                    column_letter=get_column_letter(cell.column),
                    value=_serializable_value(cell.value, value_type),
                    value_type=value_type,
                    is_formula=value_type == "formula",
                    style=_style_profile(cell),
                    merged_range=merged_anchor.get(cell.coordinate),
                )
            )

    content_rows = {cell.row for cell in cells}
    content_columns = {cell.column for cell in cells}
    if cells:
        min_row, max_row = min(content_rows), max(content_rows)
        min_column, max_column = min(content_columns), max(content_columns)
        content_range = (
            f"{get_column_letter(min_column)}{min_row}:"
            f"{get_column_letter(max_column)}{max_row}"
        )
    else:
        min_row = max_row = min_column = max_column = None
        content_range = None

    by_row: dict[int, list[CellProfile]] = {}
    by_column: dict[int, list[CellProfile]] = {}
    for cell in cells:
        by_row.setdefault(cell.row, []).append(cell)
        by_column.setdefault(cell.column, []).append(cell)

    row_profiles: list[RowProfile] = []
    column_profiles: list[ColumnProfile] = []
    if min_row is not None and max_row is not None:
        for row_index in range(min_row, max_row + 1):
            row_cells = by_row.get(row_index, [])
            count, text, numeric, boolean, formula = _counts(row_cells)
            columns = [cell.column for cell in row_cells]
            row_profiles.append(
                RowProfile(
                    row_index=row_index,
                    non_empty_count=count,
                    text_count=text,
                    numeric_count=numeric,
                    boolean_count=boolean,
                    formula_count=formula,
                    first_non_empty_column=min(columns) if columns else None,
                    last_non_empty_column=max(columns) if columns else None,
                    is_blank=count == 0,
                )
            )
    if min_column is not None and max_column is not None:
        for column_index in range(min_column, max_column + 1):
            column_cells = by_column.get(column_index, [])
            count, text, numeric, boolean, formula = _counts(column_cells)
            rows = [cell.row for cell in column_cells]
            column_profiles.append(
                ColumnProfile(
                    column_index=column_index,
                    column_letter=get_column_letter(column_index),
                    non_empty_count=count,
                    text_count=text,
                    numeric_count=numeric,
                    boolean_count=boolean,
                    formula_count=formula,
                    first_non_empty_row=min(rows) if rows else None,
                    last_non_empty_row=max(rows) if rows else None,
                    is_blank=count == 0,
                )
            )

    freeze = worksheet.freeze_panes
    freeze_panes = getattr(freeze, "coordinate", None) if freeze is not None else None
    if freeze is not None and freeze_panes is None:
        freeze_panes = str(freeze)

    return SheetProfile(
        sheet_name=worksheet.title,
        sheet_state=worksheet.sheet_state,
        openpyxl_max_row=worksheet.max_row,
        openpyxl_max_column=worksheet.max_column,
        content_min_row=min_row,
        content_max_row=max_row,
        content_min_column=min_column,
        content_max_column=max_column,
        content_range=content_range,
        non_empty_cell_count=len(cells),
        merged_ranges=merged_ranges,
        hidden_rows=sorted(
            index for index, dimension in worksheet.row_dimensions.items() if dimension.hidden
        ),
        hidden_columns=_hidden_columns(worksheet),
        freeze_panes=freeze_panes,
        row_profiles=row_profiles,
        column_profiles=column_profiles,
        cells=cells,
        blank_row_runs=(
            _blank_runs(min_row, max_row, content_rows) if min_row is not None and max_row is not None else []
        ),
        blank_column_runs=(
            _blank_runs(min_column, max_column, content_columns)
            if min_column is not None and max_column is not None
            else []
        ),
        candidate_regions=_candidate_regions(cells),
    )


def profile_workbook(path: Path | str) -> WorkbookProfile:
    """Open and deterministically profile every worksheet in workbook order."""
    source_path = Path(path)
    source_hash = source_file_sha256(source_path)
    workbook = openpyxl.load_workbook(source_path, data_only=False, read_only=False)
    try:
        active_sheet = workbook.active.title if workbook.worksheets else None
        return WorkbookProfile(
            source_file_name=source_path.name,
            source_file_sha256=source_hash,
            workbook_sheet_names=list(workbook.sheetnames),
            sheets=[profile_sheet(sheet) for sheet in workbook.worksheets],
            defined_names_count=len(workbook.defined_names),
            active_sheet=active_sheet,
        )
    finally:
        workbook.close()
